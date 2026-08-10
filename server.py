#!/usr/bin/env python3
from __future__ import annotations

import json
import re
import io
import html
import os
import random
import shutil
import time
import urllib.parse
import urllib.request
from datetime import date
from html.parser import HTMLParser
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from pypdf import PdfReader

try:
    import fitz  # PyMuPDF
except Exception:  # pragma: no cover - optional stronger PDF parser.
    fitz = None

try:
    from pdfminer.high_level import extract_text as pdfminer_extract_text
except Exception:  # pragma: no cover - optional stronger PDF parser.
    pdfminer_extract_text = None

try:
    from PIL import Image
except Exception:  # pragma: no cover - optional OCR dependency.
    Image = None

try:
    import pytesseract
except Exception:  # pragma: no cover - optional OCR dependency.
    pytesseract = None

try:
    from scrapling.fetchers import Fetcher as ScraplingFetcher
except Exception:  # pragma: no cover - Scrapling is an optional local crawler dependency.
    ScraplingFetcher = None

ROOT = Path(__file__).resolve().parent
HOST = "0.0.0.0"
PORT = int(os.environ.get("PORT", "8765"))
CURRENT_YEAR = date.today().year
CURRENT_ROC_YEAR = CURRENT_YEAR - 1911
DEFAULT_COMPARE_ROC_YEARS = [CURRENT_ROC_YEAR - offset for offset in range(3)]
CACHE_TTL_SECONDS = int(os.environ.get("CACHE_TTL_SECONDS", "1800"))
APP_VERSION = os.environ.get("RENDER_GIT_COMMIT", "dev")
CACHE_DIR = ROOT / ".cache"
NOTICE_CACHE_PATH = CACHE_DIR / "mops_notice_cache.json"
NOTICE_SEED_CACHE_PATH = ROOT / "data" / "mops_notice_seed_cache.json"
MOPS_ATTEMPT_LOG_PATH = ROOT / "data" / "mops_notice_attempts.json"
OFFICIAL_SITE_SCAN_CACHE_PATH = ROOT / "data" / "official_site_scan_cache.json"
LOOKUP_SNAPSHOT_PATH = ROOT / "data" / "lookup_snapshot.json"
REQUESTED_CODES_PATH = CACHE_DIR / "requested_codes.json"
NOTICE_PDF_DIR = CACHE_DIR / "mops-notices"
NOTICE_CACHE_VERSION = 7
EXCEL_ILLEGAL_CHAR_RE = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")

WESPAI_URL = f"https://stock.wespai.com/stock{CURRENT_ROC_YEAR}"
IDEAL_URL = "https://souvenir.ideal-labs.com/"
HONSEC_URL = "https://srd.honsec.com.tw/stock/souvenir.aspx"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,application/pdf;q=0.8,*/*;q=0.7",
    "Accept-Language": "zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7",
    "Cache-Control": "no-cache",
    "Pragma": "no-cache",
    "Upgrade-Insecure-Requests": "1",
}
REQUEST_DELAY_MIN_MS = int(os.environ.get("HTTP_REQUEST_DELAY_MIN_MS", "0"))
REQUEST_DELAY_MAX_MS = int(os.environ.get("HTTP_REQUEST_DELAY_MAX_MS", "0"))
ALLOW_LIVE_LOOKUP = os.environ.get("ALLOW_LIVE_LOOKUP", "").lower() in {"1", "true", "yes", "on"}
MOPS_FETCH_ENGINE = os.environ.get("MOPS_FETCH_ENGINE", "auto").lower()
PDF_OCR_MODE = os.environ.get("PDF_OCR_MODE", "auto").lower()
PDF_OCR_LANG = os.environ.get("PDF_OCR_LANG", "chi_tra+eng")
PDF_OCR_MAX_PAGES = int(os.environ.get("PDF_OCR_MAX_PAGES", "3"))
PDF_OCR_MIN_SCORE = int(os.environ.get("PDF_OCR_MIN_SCORE", "80"))

CACHE: dict[str, tuple[float, Any]] = {}
NOTICE_CACHE_MEMORY: dict[str, Any] | None = None
LOOKUP_SNAPSHOT_MEMORY: dict[int, dict[str, Any]] = {}

CACHE_DIR.mkdir(exist_ok=True)
NOTICE_PDF_DIR.mkdir(exist_ok=True)


def normalize_roc_year(value: int | str | None = None) -> int:
    if value in {None, ""}:
        return CURRENT_ROC_YEAR
    raw = str(value).strip()
    if not raw:
        return CURRENT_ROC_YEAR
    match = re.search(r"\d{3,4}", raw)
    if not match:
        raise ValueError(f"年度格式錯誤：{value}")
    year = int(match.group(0))
    if year >= 1912:
        year -= 1911
    if year < 90 or year > CURRENT_ROC_YEAR:
        raise ValueError(f"年度超出可用範圍：{value}")
    return year


def parse_roc_years(raw: str | None) -> list[int]:
    if not raw:
        return [CURRENT_ROC_YEAR]
    years: list[int] = []
    for token in re.findall(r"\d{3,4}", raw):
        roc_year = normalize_roc_year(token)
        if roc_year not in years:
            years.append(roc_year)
    return years or [CURRENT_ROC_YEAR]


def wespai_url_for_year(roc_year: int | str | None = None) -> str:
    return f"https://stock.wespai.com/stock{normalize_roc_year(roc_year)}"


def notice_cache_storage_key(code: str, roc_year: int | str | None = None) -> str:
    normalized_year = normalize_roc_year(roc_year)
    return code if normalized_year == CURRENT_ROC_YEAR else f"{normalized_year}:{code}"


def notice_cache_lookup(cache: dict[str, Any], code: str, roc_year: int | str | None = None) -> dict[str, Any] | None:
    normalized_year = normalize_roc_year(roc_year)
    keys = [notice_cache_storage_key(code, normalized_year)]
    if normalized_year == CURRENT_ROC_YEAR:
        keys.append(f"{normalized_year}:{code}")
    for key in keys:
        entry = cache.get(key)
        if isinstance(entry, dict):
            return entry
    return None


def lookup_snapshot_path_for_year(roc_year: int | str | None = None) -> Path:
    normalized_year = normalize_roc_year(roc_year)
    if normalized_year == CURRENT_ROC_YEAR:
        return LOOKUP_SNAPSHOT_PATH
    return ROOT / "data" / "lookup_snapshots" / f"lookup_snapshot_{normalized_year}.json"


def normalize_text(value: str) -> str:
    return " ".join(value.replace("\xa0", " ").split())


def clean_pickup_documents_text(value: str) -> str:
    cleaned = normalize_text(value or "").strip("：:，,。 ")
    if not cleaned:
        return ""
    cleaned = re.sub(r"\s*[，,]?\s*(?:自|於)\d{2,3}年.*$", "", cleaned)
    cleaned = re.sub(r"\s*[，,]?\s*(?:自|於)\d{1,3}/\d{1,2}/\d{1,2}.*$", "", cleaned)
    cleaned = re.sub(r"\s*(?:自|於)\d{2,3}年.*$", "", cleaned)
    cleaned = re.sub(r"\s*(?:自|於)\d{1,3}/\d{1,2}/\d{1,2}.*$", "", cleaned)
    cleaned = cleaned.strip("：:，,。 ")
    for left, right in (("(", ")"), ("（", "）"), ("「", "」"), ("『", "』")):
        diff = cleaned.count(left) - cleaned.count(right)
        if 0 < diff <= 2:
            cleaned += right * diff
    return cleaned


def strip_notice_noise(text: str) -> str:
    cleaned = compact_text(text or "")
    if not cleaned:
        return ""

    stop_patterns = [
        r"window\.focus\(\)",
        r"背面\s*\d*\.indd",
        r"簽名或蓋章領取日期",
        r"領取日期年月日",
        r"股務代理人股務代理部收",
        r"寄件人[:：]",
        r"國內郵資",
        r"股東戶號[:：]",
        r"股東或代\s*理人姓名",
        r"持有股數[:：]",
        r"親自出席簽名或蓋章",
        r"委任股東",
        r"徵求人",
        r"受託代理人",
        r"第\s*[一二三四五六七八九十0-9]+\s*聯",
        r"委託書填表須知",
        r"※股東、徵求人、受託代理人",
        r"﹏+",
    ]
    end = len(cleaned)
    for pattern in stop_patterns:
        match = re.search(pattern, cleaned, re.I)
        if match:
            end = min(end, match.start())
    cleaned = cleaned[:end]
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.strip("：:，,。 ;；")


def trim_notice_summary(text: str) -> str:
    cleaned = strip_notice_noise(text)
    if not cleaned:
        return ""

    cut_markers = [
        "股東如欲委託代理出席領取紀念品時",
        "貴股東如不克親自出席欲委託徵求人出席股東會",
        "貴股東如欲委託徵求人出席並領取紀念品",
        "紀念品領取方式如下",
        "採電子投票之股東",
        "電子投票領取紀念品方式",
        "電子方式行使表決權且投票成功者",
        "電子投票成功且未以其他方式出席股東會之股東",
        "採電子方式行使表決權之股東",
        "以電子方式行使表決權者",
        "本次股東會如有公開徵求委託書之情事",
        "委託書用紙填發須知",
        "※洽領紀念品須知※",
    ]
    for marker in cut_markers:
        idx = cleaned.find(marker)
        if idx > 0:
            cleaned = cleaned[:idx]
            break

    cleaned = re.sub(r"[（(]持股1,000股以上[）)][:：]?", "", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.strip("：:，,。 ;；")


EVOTE_SIGNAL_PATTERN = (
    r"(?:電子(?:投票|方式行使表決權)|以電子方式行使表決權|"
    r"採電子(?:投票|方式行使表決權)|完成電子投票|投票成功|電子投票方式出席)"
)
GIFT_PICKUP_PATTERN = (
    r"(?:紀念品|贈品|議案表決情形|領取期間|領取地點|領取時間|"
    r"領取方式|領取.{0,24}(?:紀念品|贈品)|換領(?:紀念品|贈品)?|洽領(?:紀念品|贈品)?)"
)
PICKUP_EVIDENCE_RE = re.compile(
    rf"(?:{EVOTE_SIGNAL_PATTERN}.{{0,180}}{GIFT_PICKUP_PATTERN}|"
    r"此期間非以電子.{0,120}(?:恕不|不發放).{0,30}(?:紀念品|贈品))"
)
CORE_PICKUP_EVIDENCE_RE = PICKUP_EVIDENCE_RE
VOTE_PERIOD_ONLY_RE = re.compile(r"(行使期間|電子投票期間|股東e服務|電子投票平台)")


def has_evote_pickup_evidence(text: str) -> bool:
    return bool(PICKUP_EVIDENCE_RE.search(compact_text(text or "")))


def has_core_pickup_evidence(text: str) -> bool:
    return bool(CORE_PICKUP_EVIDENCE_RE.search(compact_text(text or "")))


def has_meaningful_pickup_rule(text: str) -> bool:
    compact = compact_text(text or "")
    if not compact:
        return False
    if re.search(r"(無發放紀念品|不發放紀念品|無發放股東會紀念品)", compact) and not re.search(
        r"(領取|換領|發放期間|領取期間|攜帶|請持|議案表決情形|完成電子投票|投票成功)",
        compact,
    ):
        return False
    return has_evote_pickup_evidence(compact)


def looks_like_vote_period_only(text: str) -> bool:
    compact = compact_text(text or "")
    if not compact:
        return False
    return bool(VOTE_PERIOD_ONLY_RE.search(compact)) and not has_evote_pickup_evidence(compact)


def json_response(handler: SimpleHTTPRequestHandler, payload: dict[str, Any], status: int = 200) -> None:
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.send_header("Cache-Control", "no-store")
    handler.end_headers()
    handler.wfile.write(body)


def polite_request_delay() -> None:
    delay_min = max(0, REQUEST_DELAY_MIN_MS)
    delay_max = max(delay_min, REQUEST_DELAY_MAX_MS)
    if delay_max <= 0:
        return
    time.sleep(random.uniform(delay_min, delay_max) / 1000)


def request_headers(extra_headers: dict[str, str] | None = None) -> dict[str, str]:
    headers = dict(HEADERS)
    if extra_headers:
        headers.update(extra_headers)
    return headers


def fetch_text(url: str, timeout: int = 30, extra_headers: dict[str, str] | None = None) -> str:
    polite_request_delay()
    request = urllib.request.Request(url, headers=request_headers(extra_headers))
    with urllib.request.urlopen(request, timeout=timeout) as response:
        charset = response.headers.get_content_charset() or "utf-8"
        return response.read().decode(charset, "ignore")


def fetch_bytes(
    url: str,
    data: bytes | None = None,
    timeout: int = 30,
    extra_headers: dict[str, str] | None = None,
) -> bytes:
    headers = request_headers(extra_headers)
    if data is not None:
        headers["Content-Type"] = "application/x-www-form-urlencoded"
    polite_request_delay()
    request = urllib.request.Request(url, data=data, headers=headers)
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return response.read()


def fetch_text_with_encoding(url: str, encoding: str, data: bytes | None = None, timeout: int = 30) -> str:
    return fetch_bytes(url, data=data, timeout=timeout).decode(encoding, "ignore")


def fetch_mops_text_with_encoding(url: str, encoding: str = "big5", timeout: int = 30) -> str:
    if MOPS_FETCH_ENGINE in {"scrapling", "auto"} and ScraplingFetcher is not None:
        try:
            return fetch_text_with_scrapling(url, timeout=timeout)
        except Exception as error:
            if MOPS_FETCH_ENGINE == "scrapling":
                raise RuntimeError(f"Scrapling MOPS fetch failed: {error}") from error
    return fetch_text_with_encoding(url, encoding, timeout=timeout)


def fetch_text_with_scrapling(url: str, timeout: int = 30) -> str:
    if ScraplingFetcher is None:
        raise RuntimeError("Scrapling is not installed")

    polite_request_delay()
    attempts = (
        {"timeout": timeout, "follow_redirects": True, "stealthy_headers": True},
        {"timeout": timeout, "follow_redirects": True},
    )
    last_error: Exception | None = None
    for kwargs in attempts:
        try:
            page = ScraplingFetcher.get(url, **kwargs)
            for attr in ("html_content", "body", "text", "html"):
                value = getattr(page, attr, "")
                if callable(value):
                    value = value()
                if isinstance(value, str):
                    text = value.strip()
                    if text and text.lower() != "none":
                        return value
            rendered = str(page)
            if rendered.strip() and rendered.strip().lower() != "none":
                return rendered
        except TypeError as error:
            last_error = error
        except Exception as error:
            last_error = error
    raise RuntimeError(str(last_error or "Scrapling could not fetch page"))


def cached(name: str, loader) -> Any:
    now = time.time()
    item = CACHE.get(name)
    if item and now - item[0] < CACHE_TTL_SECONDS:
        return item[1]
    value = loader()
    CACHE[name] = (now, value)
    return value


def safe_cached_source(name: str, loader) -> dict[str, Any]:
    try:
        value = cached(name, loader)
        return value if isinstance(value, dict) else {}
    except Exception as error:
        print(f"[source:{name}] load failed: {error}")
        CACHE[name] = (time.time(), {})
        return {}


def load_notice_cache() -> dict[str, Any]:
    global NOTICE_CACHE_MEMORY
    if NOTICE_CACHE_MEMORY is not None:
        return NOTICE_CACHE_MEMORY
    seed_cache: dict[str, Any] = {}
    if NOTICE_SEED_CACHE_PATH.exists():
        try:
            seed_cache = json.loads(NOTICE_SEED_CACHE_PATH.read_text(encoding="utf-8"))
        except Exception:
            seed_cache = {}
    if NOTICE_CACHE_PATH.exists():
        try:
            disk_cache = json.loads(NOTICE_CACHE_PATH.read_text(encoding="utf-8"))
            # The deployed seed cache is curated and versioned, so it must win over
            # stale runtime cache entries that may have been created before a fix.
            NOTICE_CACHE_MEMORY = {**disk_cache, **seed_cache}
            return NOTICE_CACHE_MEMORY
        except Exception:
            pass
    NOTICE_CACHE_MEMORY = seed_cache
    return NOTICE_CACHE_MEMORY


def load_lookup_snapshot(roc_year: int | str | None = None) -> dict[str, Any]:
    global LOOKUP_SNAPSHOT_MEMORY
    normalized_year = normalize_roc_year(roc_year)
    if normalized_year in LOOKUP_SNAPSHOT_MEMORY:
        return LOOKUP_SNAPSHOT_MEMORY[normalized_year]
    snapshot_path = lookup_snapshot_path_for_year(normalized_year)
    if snapshot_path.exists():
        try:
            payload = json.loads(snapshot_path.read_text(encoding="utf-8"))
            if isinstance(payload, dict):
                LOOKUP_SNAPSHOT_MEMORY[normalized_year] = payload
                return LOOKUP_SNAPSHOT_MEMORY[normalized_year]
        except Exception:
            pass
    LOOKUP_SNAPSHOT_MEMORY[normalized_year] = {}
    return LOOKUP_SNAPSHOT_MEMORY[normalized_year]


def load_json_dict(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return payload if isinstance(payload, dict) else {}


def save_notice_cache(cache: dict[str, Any]) -> None:
    global NOTICE_CACHE_MEMORY
    NOTICE_CACHE_MEMORY = cache
    NOTICE_CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def build_notice_progress() -> dict[str, Any]:
    watchlist_path = ROOT / "data" / "mops_seed_watchlist.txt"
    watchlist_codes = clean_codes(watchlist_path.read_text(encoding="utf-8")) if watchlist_path.exists() else []
    watchlist_set = set(watchlist_codes)
    cache = load_notice_cache()
    attempt_log = load_json_dict(MOPS_ATTEMPT_LOG_PATH)
    official_scan_cache: dict[str, Any] = {}
    if OFFICIAL_SITE_SCAN_CACHE_PATH.exists():
        try:
            official_scan_cache = json.loads(OFFICIAL_SITE_SCAN_CACHE_PATH.read_text(encoding="utf-8"))
        except Exception:
            official_scan_cache = {}
    watched_cache = {code: cache.get(code) for code in watchlist_codes if cache.get(code)}
    with_notice = [code for code, item in watched_cache.items() if item.get("filename") or item.get("pdfUrl")]
    with_pickup_date = [
        code
        for code, item in watched_cache.items()
        if item.get("evotePickupStartDate") or item.get("evotePickupEndDate") or item.get("evotePickupPeriodText")
    ]
    company_pdf = [
        code
        for code, item in watched_cache.items()
        if item.get("sourceType") in {"company_pdf", "official_pdf", "transfer_agent_pdf"}
    ]
    tracked_attempts = {
        code: item
        for code, item in attempt_log.items()
        if code in watchlist_set and isinstance(item, dict) and item.get("attemptedAt")
    }
    # Entries already in the curated notice cache predate the attempt log, so
    # count them as inferred attempts without pretending their failure history is known.
    inferred_attempted = set(tracked_attempts) | set(with_notice)
    rate_limited_attempts = [
        code
        for code, item in tracked_attempts.items()
        if item.get("status") == "rate_limited" or "查詢過量" in str(item.get("error", ""))
    ]
    latest_fetched = max((item.get("fetchedAt", "") for item in watched_cache.values()), default="")
    latest_mops_attempt = max((item.get("attemptedAt", "") for item in tracked_attempts.values()), default="")
    official_scanned = [code for code in watchlist_codes if (official_scan_cache.get(code) or {}).get("attemptedAt")]
    official_found = [code for code in watchlist_codes if (official_scan_cache.get(code) or {}).get("foundUseful")]
    latest_official_scan = max(((official_scan_cache.get(code) or {}).get("attemptedAt", "") for code in watchlist_codes), default="")
    return {
        "watchlistTotal": len(watchlist_codes),
        "noticeCached": len(with_notice),
        "pickupDateCached": len(with_pickup_date),
        "officialPdfCached": len(company_pdf),
        "mopsAttemptLogged": len(tracked_attempts),
        "mopsAttemptedOrCached": len(inferred_attempted),
        "mopsNeverAttempted": max(0, len(watchlist_codes) - len(inferred_attempted)),
        "mopsRateLimited": len(rate_limited_attempts),
        "officialSiteScanned": len(official_scanned),
        "officialSiteFound": len(official_found),
        "missingNotice": max(0, len(watchlist_codes) - len(with_notice)),
        "missingPickupDate": max(0, len(watchlist_codes) - len(with_pickup_date)),
        "latestFetchedAt": latest_fetched,
        "latestMopsAttemptAt": latest_mops_attempt,
        "latestOfficialSiteScan": latest_official_scan,
    }


def build_source_stats(sources: dict[str, Any]) -> dict[str, int]:
    return {
        "wespai": len(sources["wespai"]),
        "idealLabs": len(sources["ideal"]),
        "honsec": len(sources["honsec"]),
    }


def snapshot_metadata(roc_year: int | str | None = None) -> dict[str, Any]:
    normalized_year = normalize_roc_year(roc_year)
    snapshot = load_lookup_snapshot(normalized_year)
    records = snapshot.get("records") if isinstance(snapshot.get("records"), dict) else {}
    return {
        "rocYear": normalized_year,
        "year": normalized_year + 1911,
        "dataMode": "snapshot" if records else "live",
        "snapshotGeneratedAt": snapshot.get("generatedAt", "") if records else "",
        "snapshotRecordCount": len(records),
    }


def load_requested_codes() -> dict[str, str]:
    if REQUESTED_CODES_PATH.exists():
        try:
            return json.loads(REQUESTED_CODES_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def record_requested_codes(codes: list[str]) -> None:
    requested = load_requested_codes()
    now = time.strftime("%Y-%m-%dT%H:%M:%S%z")
    changed = False
    for code in codes:
        if re.fullmatch(r"\d{3,6}", code):
            requested[code] = now
            changed = True
    if changed:
        REQUESTED_CODES_PATH.write_text(json.dumps(requested, ensure_ascii=False, indent=2), encoding="utf-8")


def compact_text(value: str) -> str:
    return re.sub(r"\s+", "", value)


def parse_compact_roc_date(value: str, fallback_year: int | None = None) -> str | None:
    match = re.search(r"(?:(\d{2,3})年)?(\d{1,2})月(\d{1,2})日", value)
    if not match:
        return None
    roc_year = int(match.group(1)) if match.group(1) else fallback_year
    if roc_year is None:
        return None
    western_year = roc_year + 1911
    return f"{western_year:04d}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"


def parse_flexible_roc_date(value: str, fallback_year: int | None = None) -> str | None:
    compact = compact_text(value).replace("民國", "")
    match = re.search(r"(?:(\d{2,3})年)?(\d{1,2})月(\d{1,2})日", compact)
    if match:
        roc_year = int(match.group(1)) if match.group(1) else fallback_year
        if roc_year is None:
            return None
        return f"{roc_year + 1911:04d}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"

    slash_match = re.search(r"(?:(\d{2,3})/)?(\d{1,2})/(\d{1,2})", compact)
    if slash_match:
        roc_year = int(slash_match.group(1)) if slash_match.group(1) else fallback_year
        if roc_year is None:
            return None
        return f"{roc_year + 1911:04d}-{int(slash_match.group(2)):02d}-{int(slash_match.group(3)):02d}"
    return None


def parse_compact_roc_range_from_text(value: str) -> tuple[str | None, str | None, str]:
    compact = compact_text(value)
    patterns = [
        r"(自?((\d{2,3})年\d{1,2}月\d{1,2}日)起至((?:\d{2,3}年)?\d{1,2}月\d{1,2}日)止)",
        r"(自?((\d{2,3})年\d{1,2}月\d{1,2}日)至((?:\d{2,3}年)?\d{1,2}月\d{1,2}日))",
        r"(自?((\d{2,3})年\d{1,2}月\d{1,2}日)起至(\d{1,2}月\d{1,2}日)止)",
        r"(自?((\d{2,3})年\d{1,2}月\d{1,2}日)至(\d{1,2}月\d{1,2}日))",
    ]
    for pattern in patterns:
        match = re.search(pattern, compact)
        if not match:
            continue
        start_date = parse_compact_roc_date(match.group(2))
        end_date = parse_compact_roc_date(match.group(4), fallback_year=int(match.group(3)))
        return start_date, end_date, match.group(1)

    single_match = re.search(r"((\d{2,3})年\d{1,2}月\d{1,2}日)", compact)
    if single_match:
        single_date = parse_compact_roc_date(single_match.group(1))
        return single_date, single_date, single_match.group(1)

    return None, None, ""


def parse_pickup_roc_range_from_text(value: str) -> tuple[str | None, str | None, str]:
    compact = compact_text(value)
    normalized = compact.replace("民國", "")
    separator = r"(?:至|[-~～—]+)"
    candidates: list[tuple[int, re.Match[str]]] = []
    patterns = [
        (0, rf"(請於((\d{{2,3}})年\d{{1,2}}月\d{{1,2}}日)(?:起)?{separator}((?:\d{{2,3}}年)?\d{{1,2}}月\d{{1,2}}日)(?:止)?)"),
        (0, rf"(請?自((\d{{2,3}})年\d{{1,2}}月\d{{1,2}}日)(?:起)?{separator}((?:\d{{2,3}}年)?\d{{1,2}}月\d{{1,2}}日)(?:止)?)"),
        (0, rf"(領取紀念品期間[:：]?((\d{{2,3}})年\d{{1,2}}月\d{{1,2}}日)(?:起)?{separator}((?:\d{{2,3}}年)?\d{{1,2}}月\d{{1,2}}日)(?:止)?)"),
        (1, rf"(領取期間及地點[:：]?自?((\d{{2,3}})年\d{{1,2}}月\d{{1,2}}日)(?:起)?{separator}((?:\d{{2,3}}年)?\d{{1,2}}月\d{{1,2}}日)(?:止)?)"),
        (2, rf"(於((\d{{2,3}})年\d{{1,2}}月\d{{1,2}}日)(?:起)?{separator}((?:\d{{2,3}}年)?\d{{1,2}}月\d{{1,2}}日)(?:止)?)"),
        (0, rf"(((\d{{2,3}})/\d{{1,2}}/\d{{1,2}})\s*{separator}\s*((?:\d{{2,3}}/)?\d{{1,2}}/\d{{1,2}}))"),
    ]
    for priority, pattern in patterns:
        for match in re.finditer(pattern, normalized):
            preceding = normalized[max(0, match.start() - 40) : match.start()]
            if re.search(r"(行使期間為?|電子投票期間[:：為]?)", preceding):
                continue
            following = normalized[match.end() : match.end() + 140]
            if priority == 2 and not re.search(r"(領取|換領|發放|股務|紀念品)", following):
                continue
            candidates.append((priority, match))
    if candidates:
        def candidate_score(item: tuple[int, re.Match[str]]) -> tuple[int, int, int]:
            priority, match = item
            before = normalized[max(0, match.start() - 140) : match.start()]
            after = normalized[match.end() : match.end() + 260]
            context = before + match.group(1) + after
            score = 100 - priority * 10
            if re.search(r"(領取|換領|領獎|發放|洽領)", after):
                score += 35
            if re.search(r"(紀念品|股務代理部|股代|地址|地點)", after):
                score += 18
            if re.search(r"(身分證|戶口名簿|出席通知書|議案表決情形|上述.*文件|請持|攜帶|憑)", before):
                score += 22
            if re.search(r"(此期間非|逾時不發放|逾期不補發|電子方式行使表決權之股東)", context):
                score += 18
            if re.search(r"(委託書徵求期間|徵求期間|徵求人|證基會|free\.sfi)", context):
                score -= 35
            if re.search(r"(行使期間為|股東e服務|電子投票平台)", context) and not re.search(r"(領取|換領|紀念品)", context):
                score -= 45
            if re.search(r"(股東會當日|開會當天會議結束前|會場領取)", context) and not re.search(r"(電子方式|電子投票|投票成功)", context):
                score -= 15
            return (-score, priority, match.start())

        priority, match = sorted(candidates, key=candidate_score)[0]
        context = normalized[max(0, match.start() - 120) : match.end() + 260]
        if looks_like_vote_period_only(context) or not has_evote_pickup_evidence(context):
            return None, None, ""
        start_date = parse_flexible_roc_date(match.group(2))
        end_date = parse_flexible_roc_date(match.group(4), fallback_year=int(match.group(3)))
        return start_date, end_date, match.group(1)

    slash_match = re.search(r"((\d{2,3})/\d{1,2}/\d{1,2})\s*(?:至|[-~～—]+)\s*((?:\d{2,3}/)?\d{1,2}/\d{1,2})", normalized)
    if slash_match:
        start_date = parse_flexible_roc_date(slash_match.group(1))
        end_date = parse_flexible_roc_date(slash_match.group(3), fallback_year=int(slash_match.group(2)))
        return start_date, end_date, slash_match.group(0)
    return None, None, ""


def parse_mmdd_to_iso(value: str, western_year: int | None = None) -> str | None:
    cleaned = normalize_text(value)
    match = re.fullmatch(r"(\d{2})\.(\d{2})", cleaned)
    if not match:
        return None
    month, day = match.groups()
    return f"{western_year or CURRENT_YEAR}-{month}-{day}"


def parse_roc_date(value: str) -> str | None:
    cleaned = normalize_text(value)
    match = re.fullmatch(r"(\d{2,3})/(\d{1,2})/(\d{1,2})", cleaned)
    if not match:
        return None
    roc_year, month, day = match.groups()
    western_year = int(roc_year) + 1911
    return f"{western_year:04d}-{int(month):02d}-{int(day):02d}"


def parse_roc_range(value: str) -> tuple[str | None, str | None]:
    cleaned = normalize_text(value)
    if cleaned in {"", "-"}:
        return None, None
    normalized = cleaned.replace("～", "~").replace("—", "-")
    match = re.fullmatch(r"(\d{2,3}/\d{1,2}/\d{1,2})\s*[-~]\s*((?:\d{2,3}/)?\d{1,2}/\d{1,2})", normalized)
    if not match:
        single = parse_roc_date(cleaned)
        return single, single
    start = parse_roc_date(match.group(1))
    second = match.group(2)
    if re.fullmatch(r"\d{1,2}/\d{1,2}", second):
        second = f"{match.group(1).split('/')[0]}/{second}"
    return start, parse_roc_date(second)


class TableCell:
    def __init__(self, text: str) -> None:
        self.text = text


class SimpleTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.tables: list[list[list[TableCell]]] = []
        self._stack: list[dict[str, Any]] = []
        self._in_cell = False
        self._buffer: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag == "table":
            self._stack.append({"rows": [], "row": None})
        elif tag == "tr" and self._stack:
            self._stack[-1]["row"] = []
        elif tag in {"td", "th"} and self._stack and self._stack[-1]["row"] is not None:
            self._in_cell = True
            self._buffer = []

    def handle_data(self, data: str) -> None:
        if self._in_cell:
            self._buffer.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag in {"td", "th"} and self._in_cell:
            text = normalize_text(" ".join(self._buffer))
            self._stack[-1]["row"].append(TableCell(text))
            self._in_cell = False
        elif tag == "tr" and self._stack and self._stack[-1]["row"] is not None:
            self._stack[-1]["rows"].append(self._stack[-1]["row"])
            self._stack[-1]["row"] = None
        elif tag == "table" and self._stack:
            self.tables.append(self._stack.pop()["rows"])


class VisibleTextParser(HTMLParser):
    SKIP_TAGS = {"script", "style", "noscript"}

    def __init__(self) -> None:
        super().__init__()
        self.items: list[str] = []
        self._skip_depth = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag in self.SKIP_TAGS:
            self._skip_depth += 1

    def handle_endtag(self, tag: str) -> None:
        if tag in self.SKIP_TAGS and self._skip_depth:
            self._skip_depth -= 1

    def handle_data(self, data: str) -> None:
        if self._skip_depth:
            return
        text = normalize_text(data)
        if text:
            self.items.append(text)


def load_wespai_for_year(roc_year: int | str | None = None) -> dict[str, dict[str, Any]]:
    normalized_year = normalize_roc_year(roc_year)
    western_year = normalized_year + 1911
    source_url = wespai_url_for_year(normalized_year)
    html = fetch_text(source_url)
    parser = SimpleTableParser()
    parser.feed(html)
    target_table: list[list[TableCell]] | None = None
    for table in parser.tables:
        if not table:
            continue
        header_text = " ".join(cell.text for cell in table[0])
        if "代號" in header_text and "紀念品" in header_text and "最後買進日" in header_text:
            target_table = table
            break

    if not target_table or len(target_table) < 2:
        raise ValueError("找不到撿股讚的股東會紀念品表格")

    rows: dict[str, dict[str, Any]] = {}
    for row in target_table[1:]:
        values = [cell.text for cell in row]
        if len(values) < 15:
            continue
        code = values[1]
        if not re.fullmatch(r"\d{3,6}", code):
            continue
        rows[code] = {
            "code": code,
            "roc_year": normalized_year,
            "year": western_year,
            "company_name": values[2],
            "price_text": values[3],
            "souvenir_name": values[4],
            "meeting_date_text": values[6],
            "meeting_date": parse_mmdd_to_iso(values[6], western_year),
            "meeting_city": values[7],
            "last_buy_date_text": values[8],
            "last_buy_date": parse_mmdd_to_iso(values[8], western_year),
            "transfer_agent_short": values[9],
            "transfer_agent_phone": values[10],
            "odd_lot_mail": values[12],
            "reelection": values[13],
            "source_url": source_url,
            "official_doc_url": (
                f"https://doc.twse.com.tw/server-java/t57sb01?step=1&colorchg=1&co_id={code}"
                f"&year={normalized_year}&mtype=F&"
            ),
        }
    return rows


def load_wespai() -> dict[str, dict[str, Any]]:
    return load_wespai_for_year(CURRENT_ROC_YEAR)


def load_ideal() -> dict[str, dict[str, Any]]:
    html = fetch_text(IDEAL_URL)
    match = re.search(r'self\.__next_f\.push\(\[1,"(1c:.*?)"\]\)</script>', html)
    if not match:
        raise ValueError("找不到股東禮簿資料")
    decoded = json.loads(f"\"{match.group(1)}\"")
    payload = json.loads(decoded[3:])
    meetings = payload[3]["allMeetings"]
    return {item["stock_code"]: item for item in meetings}


def is_company_start(tokens: list[str], index: int) -> bool:
    return (
        index + 2 < len(tokens)
        and tokens[index + 1].endswith("年")
        and tokens[index + 2].startswith("代號 ")
    )


def load_honsec() -> dict[str, dict[str, Any]]:
    html = fetch_text(HONSEC_URL)
    parser = VisibleTextParser()
    parser.feed(html)
    tokens = parser.items

    schema = [
        ("distribution_rule", "發放原則"),
        ("last_buy_date_text", "最後買進日"),
        ("agent_distribution_period_text", "股代代發期間"),
        ("meeting_distribution", "股東會現場紀念品有無發放"),
        ("agent_only_distribution", "股代有無純代發紀念品"),
        ("souvenir_name", "紀念品"),
        ("meeting_distribution_rule", "開會現場發放條件"),
        ("has_evote", "是否有電子投票"),
        ("evote_period_text", "電子投票期間"),
        ("evote_pickup_place", "電投發放地點"),
        ("evote_pickup_period_text", "電投紀念品發放期間"),
        ("evote_pickup_rule", "電子投票發放條件"),
        ("proxy_agent", "受託代理人"),
        ("proxy_agent_locations", "徵求地點"),
        ("public_proxy", "個人公開徵求"),
        ("public_proxy_locations", "徵求地點"),
        ("agent_proxy", "受託或股代徵求"),
        ("agent_proxy_locations", "徵求地點"),
        ("proxy_distribution", "徵求場所紀念品有無發放"),
        ("proxy_period_text", "徵求期間"),
    ]
    known_labels = {label for _, label in schema}
    rows: dict[str, dict[str, Any]] = {}

    index = 0
    while index < len(tokens):
        if not is_company_start(tokens, index):
            index += 1
            continue

        company_name = tokens[index]
        code = tokens[index + 2].replace("代號", "").strip()
        index += 3

        record: dict[str, Any] = {
            "code": code,
            "company_name": company_name,
            "source_url": HONSEC_URL,
        }

        for field_name, label in schema:
            if index >= len(tokens) or tokens[index] != label:
                record[field_name] = ""
                continue

            index += 1
            if index >= len(tokens) or tokens[index] in known_labels or is_company_start(tokens, index):
                record[field_name] = ""
            else:
                record[field_name] = tokens[index]
                index += 1

        record["last_buy_date"] = parse_roc_date(record["last_buy_date_text"])
        record["evote_start_date"], record["evote_end_date"] = parse_roc_range(record["evote_period_text"])
        record["evote_pickup_start_date"], record["evote_pickup_end_date"] = parse_roc_range(
            record["evote_pickup_period_text"]
        )
        rows[code] = record

    return rows


def source_bundle(roc_year: int | str | None = None) -> dict[str, Any]:
    normalized_year = normalize_roc_year(roc_year)
    wespai_cache_name = "wespai" if normalized_year == CURRENT_ROC_YEAR else f"wespai-{normalized_year}"
    wespai = safe_cached_source(wespai_cache_name, lambda: load_wespai_for_year(normalized_year))
    if normalized_year == CURRENT_ROC_YEAR:
        ideal = safe_cached_source("ideal", load_ideal)
        honsec = safe_cached_source("honsec", load_honsec)
    else:
        # These sources only expose the active season reliably. Historical
        # comparisons use Wespai plus the year-specific MOPS notice cache.
        ideal = {}
        honsec = {}
    return {"wespai": wespai, "ideal": ideal, "honsec": honsec}


def fetch_notice_listing(code: str, roc_year: int | str | None = None) -> dict[str, str] | None:
    normalized_year = normalize_roc_year(roc_year)
    query_url = f"https://doc.twse.com.tw/server-java/t57sb01?step=1&colorchg=1&co_id={code}&year={normalized_year}&mtype=F&"
    html_text = html.unescape(fetch_mops_text_with_encoding(query_url, "big5"))

    def strip_tags(fragment: str) -> str:
        return normalize_text(re.sub(r"<[^>]+>", " ", fragment))

    rows = []
    for row_html in re.findall(r"<tr[^>]*>(.*?)</tr>", html_text, re.S | re.I):
        filename_match = re.search(
            rf"readfile2\(\s*['\"]F['\"]\s*,\s*['\"]{code}['\"]\s*,\s*['\"]([^'\"]+)['\"]\s*\)",
            row_html,
            re.I,
        )
        if not filename_match:
            continue

        cell_fragments = re.findall(r"<td[^>]*>(.*?)</td>", row_html, re.S | re.I)
        cells = [strip_tags(cell) for cell in cell_fragments]
        if not cells or cells[0] != code:
            continue

        detail = next((cell for cell in cells if "開會通知" in cell), "")
        if not detail:
            continue

        rows.append(
            {
                "year": cells[1] if len(cells) > 1 else "",
                "dataType": cells[2] if len(cells) > 2 else "",
                "meetingType": cells[4] if len(cells) > 4 else "",
                "detail": detail,
                "remark": cells[6] if len(cells) > 6 else "",
                "filename": normalize_text(filename_match.group(1)),
                "uploadedAt": cells[-1] if cells else "",
                "queryUrl": query_url,
            }
        )
    exact = next((row for row in rows if row["detail"] == "開會通知"), None)
    if exact:
        return exact
    fallback = next((row for row in rows if "開會通知" in row["detail"] and "英文版" not in row["detail"]), None)
    if fallback:
        return fallback

    filename_matches = re.findall(
        rf"readfile2\(\s*['\"]F['\"]\s*,\s*['\"]{code}['\"]\s*,\s*['\"]([^'\"]+\.pdf)['\"]\s*\)",
        html_text,
        re.I,
    )
    if not filename_matches:
        return None

    preferred_filename = next(
        (name for name in filename_matches if re.search(r"F01\.pdf$", name, re.I)),
        filename_matches[0],
    )
    return {
        "year": "",
        "dataType": "",
        "meetingType": "",
        "detail": "開會通知",
        "remark": "",
        "filename": normalize_text(preferred_filename),
        "uploadedAt": "",
        "queryUrl": query_url,
    }


def resolve_notice_pdf_url(code: str, filename: str) -> str:
    params = {
        "colorchg": "1",
        "step": "9",
        "kind": "F",
        "co_id": code,
        "filename": filename,
    }
    encoded_params = urllib.parse.urlencode(params)
    try:
        html_text = fetch_mops_text_with_encoding(
            f"https://doc.twse.com.tw/server-java/t57sb01?{encoded_params}",
            "big5",
        )
    except Exception:
        html_text = fetch_text_with_encoding(
            "https://doc.twse.com.tw/server-java/t57sb01",
            "big5",
            data=encoded_params.encode(),
        )

    html_text = html.unescape(html_text)
    match = re.search(r"href=[\"']([^\"']+\.pdf)[\"']", html_text, re.I)
    if not match:
        match = re.search(r"(/pdf/[^\"'<> ]+\.pdf)", html_text, re.I)
    if not match:
        summary = normalize_text(re.sub(r"<[^>]+>", " ", html_text))[:180]
        raise ValueError(f"找不到公開資訊觀測站通知書 PDF 下載連結：{summary}")
    return urllib.parse.urljoin("https://doc.twse.com.tw", match.group(1))


def pdf_text_quality_score(text: str) -> int:
    compact = compact_text(text or "")
    if not compact:
        return -999

    score = min(len(compact) // 120, 35)
    for keyword, weight in (
        ("電子", 10),
        ("投票", 10),
        ("領取", 12),
        ("換領", 10),
        ("紀念品", 12),
        ("身分證", 8),
        ("議案表決情形", 18),
        ("股務代理部", 8),
    ):
        if keyword in compact:
            score += weight

    start_date, _, _ = parse_pickup_roc_range_from_text(compact)
    if start_date:
        score += 60

    # pypdf sometimes decodes embedded subset fonts into Tibetan/Cyrillic/Oriya
    # lookalike characters. Penalize those outputs so PyMuPDF/pdfminer wins.
    suspicious = sum(
        1
        for char in compact
        if (
            "\u0400" <= char <= "\u052f"
            or "\u0600" <= char <= "\u06ff"
            or "\u0b00" <= char <= "\u0b7f"
            or "\u0f00" <= char <= "\u0fff"
            or char in {"ɿ", "", "", "՟", "༟"}
        )
    )
    if suspicious:
        score -= min(120, suspicious * 3)

    return score


def extract_notice_text_with_pypdf(pdf_bytes: bytes) -> str:
    reader = PdfReader(io.BytesIO(pdf_bytes))
    return "\n".join((page.extract_text() or "") for page in reader.pages)


def extract_notice_text_with_pymupdf(pdf_bytes: bytes) -> list[tuple[str, str]]:
    if fitz is None:
        return []
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    normal_text = "\n".join((page.get_text("text") or "") for page in doc)

    block_chunks: list[str] = []
    for page in doc:
        blocks = sorted(page.get_text("blocks"), key=lambda block: (block[1], block[0]))
        for block in blocks:
            if len(block) > 4 and block[4]:
                block_chunks.append(str(block[4]))
    block_text = "\n".join(block_chunks)

    candidates = [("pymupdf", normal_text)]
    if block_text and block_text != normal_text:
        candidates.append(("pymupdf-blocks", block_text))
    return candidates


def extract_notice_text_with_pdfminer(pdf_bytes: bytes) -> str:
    if pdfminer_extract_text is None:
        return ""
    return pdfminer_extract_text(io.BytesIO(pdf_bytes)) or ""


def extract_notice_text_with_ocr(pdf_bytes: bytes) -> str:
    if PDF_OCR_MODE in {"0", "false", "off", "disabled", "none"}:
        return ""
    if fitz is None or Image is None or pytesseract is None or shutil.which("tesseract") is None:
        return ""

    chunks: list[str] = []
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    page_limit = min(len(doc), max(1, PDF_OCR_MAX_PAGES))
    matrix = fitz.Matrix(2, 2)
    for page_index in range(page_limit):
        page = doc.load_page(page_index)
        pixmap = page.get_pixmap(matrix=matrix, alpha=False)
        image = Image.open(io.BytesIO(pixmap.tobytes("png")))
        try:
            chunks.append(pytesseract.image_to_string(image, lang=PDF_OCR_LANG))
        except Exception:
            # Missing OCR language packs should not break regular parsing.
            return ""
    return "\n".join(chunk for chunk in chunks if chunk)


def extract_notice_text_with_metadata(pdf_bytes: bytes) -> tuple[str, dict[str, Any]]:
    candidates: list[dict[str, Any]] = []

    def add_candidate(engine: str, text: str) -> None:
        cleaned = text or ""
        if not cleaned.strip():
            return
        candidates.append(
            {
                "engine": engine,
                "text": cleaned,
                "score": pdf_text_quality_score(cleaned),
            }
        )

    try:
        add_candidate("pypdf", extract_notice_text_with_pypdf(pdf_bytes))
    except Exception:
        pass

    for engine, text in extract_notice_text_with_pymupdf(pdf_bytes):
        add_candidate(engine, text)

    try:
        add_candidate("pdfminer", extract_notice_text_with_pdfminer(pdf_bytes))
    except Exception:
        pass

    best = max(candidates, key=lambda item: item["score"], default={"engine": "", "text": "", "score": -999})
    ocr_used = False
    if best["score"] < PDF_OCR_MIN_SCORE:
        ocr_text = extract_notice_text_with_ocr(pdf_bytes)
        if ocr_text.strip():
            ocr_used = True
            add_candidate("ocr-tesseract", ocr_text)
            if best["text"]:
                add_candidate("combined-ocr", f"{best['text']}\n{ocr_text}")
            best = max(candidates, key=lambda item: item["score"], default=best)

    metadata = {
        "engine": best.get("engine", ""),
        "score": int(best.get("score", 0)),
        "ocrUsed": ocr_used,
        "candidates": [
            {"engine": item["engine"], "score": int(item["score"])}
            for item in sorted(candidates, key=lambda item: item["score"], reverse=True)
        ],
    }
    return str(best.get("text", "")), metadata


def extract_notice_text(pdf_bytes: bytes) -> str:
    text, _ = extract_notice_text_with_metadata(pdf_bytes)
    return text


def extract_notice_evote_block(text: str) -> str:
    markers = [
        "電子方式行使表決權且投票成功者",
        "以電子方式行使表決權且投票成功",
        "採用電子投票行使表決權且投票成功",
        "有採用電子投票行使表決權且投票成功",
        "電子投票股東領取紀念品",
        "採電子方式行使表決權之股東",
        "採電子投票之股東",
        "電子投票之股東",
        "電子投票領取紀念品方式",
        "電子投票成功且未以其他方式",
        "以電子方式行使表決權者",
        "電子投票",
    ]
    compact = compact_text(text)
    candidates: list[str] = []
    for marker in markers:
        for match in re.finditer(re.escape(marker), compact):
            start = max(0, match.start() - 180)
            window = compact[start : match.start() + 2200]

            # If the marker sits inside a numbered pickup item, start closer to
            # that item so proxy/voting-period dates earlier in the notice do
            # not beat the actual electronic-vote pickup period.
            prefix = compact[start : match.start()]
            item_match = list(re.finditer(r"(?:[。；]|^)\d+[\.、]", prefix))
            if item_match:
                item_start = start + item_match[-1].start()
                window = compact[item_start : match.start() + 2200]

            stop_candidates = []
            for pattern in (
                r"開會通知書",
                r"委託書使用須知",
                r"股東戶號",
                r"徵求場所.*?表",
                r"長龍會議顧問股份有限公司",
                r"全通事務處理股份有限公司",
            ):
                stop = re.search(pattern, window[220:])
                if stop:
                    stop_candidates.append(stop.start() + 220)
            if stop_candidates:
                window = window[: min(stop_candidates)]
            candidates.append(window)

    if not candidates:
        return ""

    def score_candidate(candidate: str) -> tuple[int, int]:
        start_date, _, _ = parse_pickup_roc_range_from_text(candidate)
        score = 0
        if start_date:
            score += 40
        if re.search(r"(領取|換領|領獎|發放|洽領)", candidate):
            score += 20
        if "紀念品" in candidate:
            score += 12
        if "投票成功" in candidate:
            score += 20
        if "議案表決情形" in candidate:
            score += 18
        if re.search(r"(身分證|戶口名簿|出席通知書)", candidate):
            score += 8
        if "此期間非" in candidate:
            score += 20
        if re.search(r"(股務代理部|股代|地址|地點)", candidate):
            score += 8
        if re.search(r"(徵求人|委託書之委託人|受託代理人)", candidate) and not re.search(
            r"(投票成功|議案表決情形|此期間非)", candidate
        ):
            score -= 35
        if re.search(r"(不得領取|不予發放紀念品)", candidate) and not re.search(r"(如欲領取|領取方式|換領)", candidate):
            score -= 24
        if "行使期間為" in candidate and not re.search(r"(如欲領取|領取紀念品|換領紀念品)", candidate):
            score -= 20
        return (-score, len(candidate))

    best = sorted(candidates, key=score_candidate)[0]
    return normalize_text(best)


def extract_notice_summary(text: str) -> dict[str, Any]:
    compact = compact_text(text)
    summary_match = re.search(r"(?:紀念品領取說明|洽領紀念品領取須知|洽領紀念品須知|紀念品領取須知)(.*?)(註：|股東常會日期|委託書填表須知|第[一二三四五六七八九十]聯)", compact)
    summary = summary_match.group(1) if summary_match else ""
    evote_match = re.search(
        r"(?:採電子投票之股東，?紀念品領取方式：|電子投票領取紀念品方式：|電子方式行使表決權且投票成功者：?|電子投票成功且未以其他方式出席股東會之股東，?請依下列方式領取紀念品。?|採電子方式行使表決權之股東，?|以電子方式行使表決權者。?)"
        r"(.{0,1000}?)(註：|股東常會日期|委託書填表須知|第[一二三四五六七八九十]聯|股東戶號)",
        compact,
    )
    evote_rule = evote_match.group(1) if evote_match else ""
    evote_block = extract_notice_evote_block(text)

    block_has_specific_evote = any(
        marker in evote_block
        for marker in (
            "電子投票領取紀念品方式",
            "電子方式行使表決權且投票成功者",
            "電子投票成功且未以其他方式",
            "採電子方式行使表決權之股東",
            "以電子方式行使表決權者",
        )
    )
    if not evote_rule and evote_block:
        evote_rule = evote_block
    elif evote_block:
        block_start, _, _ = parse_pickup_roc_range_from_text(evote_block)
        rule_start, _, _ = parse_pickup_roc_range_from_text(evote_rule)
        if block_start and (block_has_specific_evote or not rule_start):
            evote_rule = evote_block

    pretty_summary = (
        summary.replace("A.", " A. ").replace("B.", " B. ").replace("C.", " C. ").replace("D.", " D. ").strip()
    )
    pretty_evote = (
        evote_rule.replace("A.", "A. ").replace("B.", " B. ").replace("C.", " C. ").replace("D.", " D. ").strip()
    )
    pretty_summary = trim_notice_summary(pretty_summary)
    pretty_evote = strip_notice_noise(pretty_evote)
    if looks_like_vote_period_only(pretty_evote) or (pretty_evote and not has_meaningful_pickup_rule(pretty_evote)):
        pretty_evote = ""

    # Summary should focus on gift/distribution notes; if OCR or regex picked
    # the electronic-vote block or proxy boilerplate, drop it from summary.
    if any(
        marker in pretty_summary
        for marker in (
            "電子投票領取紀念品方式",
            "電子方式行使表決權",
            "採電子投票之股東",
            "以電子方式行使表決權者",
        )
    ):
        pretty_summary = ""

    start_date = None
    end_date = None
    period_text = ""
    if pretty_evote:
        start_date, end_date, period_text = parse_pickup_roc_range_from_text(pretty_evote)

    return {
        "giftSummary": pretty_summary,
        "evotePickupRule": pretty_evote,
        "evotePickupStartDate": start_date,
        "evotePickupEndDate": end_date,
        "evotePickupPeriodText": period_text,
    }


def extract_pickup_location(source_hint: str, place_text: str, rule_text: str, transfer_agent_name: str) -> str:
    rule = compact_text(rule_text or "")
    place = normalize_text(place_text or "")

    if place == "不發":
        return "不發"

    def clean_location(value: str) -> str:
        cleaned = value.strip("：:，,。 ")
        cleaned = re.sub(r"^(?:\d{2,3}年)?\d{1,2}月\d{1,2}日(?:\([^)]*\))?至", "", cleaned)
        cleaned = re.sub(r"^[^，。；]{0,40}\)至", "", cleaned)
        return cleaned.strip("：:，,。 ")

    match = re.search(
        r"(?:請於|於)\d{2,3}年\d{1,2}月\d{1,2}日(?:起)?至(?:\d{2,3}年)?\d{1,2}月\d{1,2}日(?:止)?[^。；]{0,260}?[，,]至(.{2,260}?)(?:領取|換領)紀念品",
        rule,
    )
    if match:
        location = clean_location(match.group(1))
        if location:
            return location

    match = re.search(r"(?:請於|於)\d{2,3}年\d{1,2}月\d{1,2}日(?:起)?至(?:\d{2,3}年)?\d{1,2}月\d{1,2}日(?:止)?[^。；]{0,220}?至([^，。；]+?)(?:領取|換領)", rule)
    if match:
        location = clean_location(match.group(1))
        address_match = re.search(r"領取地點如下[:：](.+?)(?:※|紀念品領取時間|逾期|$)", rule)
        if address_match:
            address_text = clean_location(address_match.group(1))
            if address_text and address_text not in location:
                location = f"{location}；{address_text}"
        if location:
            return location

    match = re.search(r"領取地點如下[:：](.+?)(?:※|紀念品領取時間|逾期|$)", rule)
    if match:
        location = clean_location(match.group(1))
        if location:
            return location

    match = re.search(r"領取紀念品地點[:：](.+?)(?:。|；|$)", rule)
    if match:
        location = clean_location(match.group(1))
        if location:
            return location

    match = re.search(r"止(?:\(.*?\))?(?:[^至。；]{0,80})至([^。；]+?)(?:領取|換領)", rule)
    if match:
        location = clean_location(match.group(1))
        if location:
            return location

    match = re.search(r"發放地點：([^。；]+?)(?:持|發放時間|領取|$)", rule)
    if match:
        location = clean_location(match.group(1))
        if location:
            return location

    locations = re.findall(r"至([^至。；]+?)(?:領取|換領)", rule)
    if locations:
        location = clean_location(locations[-1])
        if location:
            return location

    if place and place not in {"公司", "股代", "代理部"}:
        return place

    if source_hint in {"宏遠股代", "開會通知書"} and transfer_agent_name:
        return transfer_agent_name

    return place or transfer_agent_name


def extract_pickup_documents(rule_text: str) -> str:
    rule = normalize_text(rule_text or "")
    compact = compact_text(rule_text or "")
    if not rule:
        return ""

    match = re.search(r"攜帶下列文件之一[:：](.+?)至下列地點", rule)
    if match:
        return clean_pickup_documents_text(match.group(1))

    match = re.search(r"(?:請)?攜帶以下文件(?:二擇一|擇一)?[：:]?(.{2,500}?)(?:，?於\d{2,3}年|於\d{2,3}年|至[^。；]{0,120}(?:領取|換領)|$)", rule)
    if match:
        return clean_pickup_documents_text(f"以下文件二擇一：{match.group(1)}")

    match = re.search(r"(?:請)?攜帶下列文件(?:辦理)?[：:]?(.{2,500}?)(?:，?並?(?:自|於)\d{2,3}年|(?:自|於)\d{2,3}年|至[^。；]{0,120}(?:領取|換領)|$)", rule)
    if match:
        return clean_pickup_documents_text(match.group(1))

    match = re.search(r"(?:攜帶文件|攜帶資料)[：:]?(.{2,220}?)(?:。|；|C\.|領取期間|發放期間|$)", rule)
    if match:
        return clean_pickup_documents_text(match.group(1))

    match = re.search(
        r"((?:憑|印出|攜帶)[^。；]{0,160}議案表決情形[^。；]{0,180}"
        r"(?:身分證|戶口名簿|健保卡|駕照|證明文件)[^。；]{0,100}?)"
        r"(?:，?領取期間|，?於\d{2,3}年|，?至[^。；]{0,120}(?:領取|換領)|。|；|$)",
        compact,
    )
    if match:
        return clean_pickup_documents_text(match.group(1))

    document_candidates: list[tuple[int, str]] = []
    for pattern in (
        r"請持有效的(.{2,220}?)(?:，?於\d{2,3}年|於\d{2,3}年|至[^。；]{0,100}(?:領取|換領)|$)",
        r"請持(.{2,220}?)(?:，?自\d{2,3}年|自\d{2,3}年|，?於\d{2,3}年|於\d{2,3}年|至[^。；]{0,100}(?:領取|換領)|$)",
        r"(?:憑|限持)(.{2,160}?)(?:至[^。；]+?(?:領取|換領)|，?於\d{2,3}年|$)",
    ):
        match = re.search(pattern, compact)
        if match:
            document_candidates.append((match.start(), match.group(1)))
    if document_candidates:
        return clean_pickup_documents_text(sorted(document_candidates, key=lambda item: item[0])[0][1])

    match = re.search(r"請於領取紀念品簽章處(.{2,220}?)(?:，?自\d{2,3}年|自\d{2,3}年|，?於\d{2,3}年|於\d{2,3}年|至[^。；]{0,100}(?:領取|換領)|$)", compact)
    if match:
        return clean_pickup_documents_text(f"於領取紀念品簽章處{match.group(1)}")

    match = re.search(r"攜帶(.{2,160}?)(?:至[^。；]+?領取|等擇一皆可|領取)", compact)
    if match:
        return clean_pickup_documents_text(match.group(1))

    if "身分證明文件" in compact and "股東會出席通知書" in compact:
        return "股東會出席通知書或身分證明文件"

    if "持相關證明文件" in compact or "持相關證明文件至" in compact:
        return "相關證明文件"

    if any(keyword in compact for keyword in ("身分證", "戶口名簿", "健保卡", "駕照", "出席通知書", "議案表決情形")):
        return clean_pickup_documents_text(rule)

    return ""


def is_useful_pickup_notice(notice: str) -> bool:
    cleaned = normalize_text(notice)
    if not cleaned or len(cleaned) < 5:
        return False
    if "紀念品兌換券" in cleaned or "至領取" in cleaned:
        return False
    if cleaned in {"恕不發放", "恕不發放紀念品"}:
        return False
    # Regexes that start at "限" can accidentally slice company names ending in
    # 有限公司, producing junk such as "限公司(...)".
    if cleaned.startswith("限公司"):
        return False
    return True


def collect_pickup_notices(text: str, max_chars: int = 80) -> list[str]:
    """Extract short, contextual reminders tied to e-vote gift pickup."""
    cleaned = normalize_text(text)
    if not cleaned:
        return []

    notices: list[str] = []
    seen = set()
    patterns = (
        rf"限[^。；]{{0,{max_chars}}}",
        rf"僅限[^。；]{{0,{max_chars}}}",
        rf"本人[^。；]{{0,{max_chars}}}",
        rf"(?:逾期|會場|現場|當日|股東會當日|開會當日|未依|未於|其他方式)[^。；]{{0,{max_chars}}}恕不[^。；]{{0,{max_chars}}}",
        rf"恕不(?:郵寄|補發)[^。；]{{0,{max_chars}}}",
    )
    for pattern in patterns:
        for match in re.findall(pattern, cleaned):
            notice = match.strip("：:，,。 ;；")
            if is_useful_pickup_notice(notice) and notice not in seen:
                notices.append(notice)
                seen.add(notice)
    return notices


def normalize_evote_rule(
    rule_text: str,
    period_text: str,
    location: str,
    documents: str,
) -> str:
    cleaned = strip_notice_noise(rule_text)
    generic_locations = {"會場", "自辦", "公司", "本公司"}
    # `strip_notice_noise` intentionally cuts noisy mailer/ticket tails, but
    # some notices place the strongest e-vote pickup evidence in that area.
    # Accept evidence from either the cleaned snippet or the parser-selected
    # original snippet, then keep the output structured via period/location/docs.
    has_explicit_pickup = has_evote_pickup_evidence(cleaned) or has_evote_pickup_evidence(rule_text)
    if not has_explicit_pickup:
        return ""
    if not period_text and not documents and location in generic_locations:
        return ""

    base_parts: list[str] = []
    if period_text:
        base_parts.append(f"領取時間：{period_text}")
    if location:
        base_parts.append(f"領取地點：{location}")
    if documents:
        base_parts.append(f"攜帶文件：{documents}")

    notices: list[str] = []
    if cleaned:
        normalized = (
            cleaned.replace("A.", "；A.")
            .replace("B.", "；B.")
            .replace("C.", "；C.")
            .replace("D.", "；D.")
            .replace("◎", "；")
        )
        parts = re.split(r"[。；]", normalized)
        include_markers = (
            "電子投票",
            "電子方式行使表決權",
            "採電子投票",
            "投票成功",
            "議案表決情形",
            "出席通知書",
            "身分證",
            "戶口名簿",
            "健保卡",
            "駕照",
            "證明文件",
            "恕不",
            "本人",
        )
        exclude_markers = (
            "委託書",
            "徵求人",
            "受託代理",
            "公司法",
            "開會二日前",
            "開會五日前",
            "簽到卡時間",
            "股東戶號",
            "股東戶名",
            "持有股數",
            "紀念品兌換券",
            "郵簡內裝有附件",
        )
        seen = set()
        for raw in parts:
            part = raw.strip("：:，,。 ;；")
            if not part or len(part) < 4:
                continue
            if any(marker in part for marker in exclude_markers):
                continue
            if (
                any(marker in part for marker in include_markers)
                or (period_text and period_text in part)
                or (location and location in part)
                or (documents and documents in part)
            ):
                for notice in collect_pickup_notices(part, 80):
                    if notice not in seen:
                        notices.append(notice)
                        seen.add(notice)

    if notices:
        base_parts.append(f"補充：{'；'.join(notices)}")

    if base_parts:
        return "；".join(base_parts)
    return ""


def compose_notice_summary(
    period_text: str,
    location: str,
    documents: str,
    rule_text: str,
    fallback_summary: str,
) -> str:
    normalized_rule = normalize_evote_rule(rule_text, period_text, location, documents)
    if not normalized_rule:
        return ""
    if not period_text and not documents:
        return ""

    parts: list[str] = []
    if period_text:
        parts.append(f"領取時間：{period_text}")
    if location:
        parts.append(f"領取地點：{location}")
    if documents:
        parts.append(f"攜帶文件：{documents}")

    cleaned_rule = normalized_rule or strip_notice_noise(rule_text)
    if cleaned_rule:
        cleaned_rule = re.sub(r"\s+", " ", cleaned_rule).strip("：:，,。 ;；")
        if period_text:
            cleaned_rule = cleaned_rule.replace(period_text, "").strip("：:，,。 ;；")
        if location:
            cleaned_rule = cleaned_rule.replace(location, "").strip("：:，,。 ;；")
        if documents:
            cleaned_rule = cleaned_rule.replace(documents, "").strip("：:，,。 ;；")
        notices = collect_pickup_notices(cleaned_rule, 60)
        if notices:
            deduped = []
            seen = set()
            for notice in notices:
                if is_useful_pickup_notice(notice) and notice not in seen:
                    deduped.append(notice)
                    seen.add(notice)
            if deduped:
                parts.append(f"補充：{'；'.join(deduped)}")

    if parts:
        return "；".join(part for part in parts if part)
    return trim_notice_summary(fallback_summary)


def merge_notice_cache_into_record(record: dict[str, Any], cached_notice: dict[str, Any] | None) -> dict[str, Any]:
    if not isinstance(cached_notice, dict):
        return record

    field_map = {
        "filename": "noticeFilename",
        "uploadedAt": "noticeUploadedAt",
        "cacheStatus": "noticeCacheStatus",
        "sourceLabel": "noticeSourceLabel",
        "sourceType": "noticeSourceType",
        "giftSummary": "noticeGiftSummary",
        "evotePickupRule": "evotePickupRule",
        "evotePickupStartDate": "evotePickupStartDate",
        "evotePickupEndDate": "evotePickupEndDate",
        "evotePickupPeriodText": "noticeEvotePickupPeriodText",
        "evotePickupLocation": "evotePickupLocation",
        "evotePickupDocuments": "evotePickupDocuments",
    }
    for cache_key, record_key in field_map.items():
        cache_value = cached_notice.get(cache_key)
        if cache_value and not record.get(record_key):
            record[record_key] = cache_value

    has_cached_pickup = bool(
        cached_notice.get("evotePickupRule")
        or cached_notice.get("evotePickupStartDate")
        or cached_notice.get("evotePickupEndDate")
        or cached_notice.get("evotePickupPeriodText")
    )
    if has_cached_pickup and not record.get("evotePickupSource"):
        record["evotePickupSource"] = cached_notice.get("sourceLabel") or "開會通知書"
    if cached_notice.get("filename") or cached_notice.get("pdfUrl"):
        record["mopsAttempted"] = True

    pdf_url = cached_notice.get("pdfUrl")
    if pdf_url:
        sources = record.get("sources")
        if not isinstance(sources, list):
            sources = []
        if not any(isinstance(source, dict) and source.get("url") == pdf_url for source in sources):
            sources.append(source_link(cached_notice.get("sourceLabel") or "開會通知書", pdf_url))
        record["sources"] = sources

    return record


def enrich_record_notice_fields(record: dict[str, Any]) -> dict[str, Any]:
    raw_evote_pickup_rule = record.get("evotePickupRule", "")
    evote_pickup_place = record.get("evotePickupPlace", "")
    transfer_agent_name = record.get("transferAgentName", "") or record.get("transferAgentShort", "")
    evote_pickup_source = record.get("evotePickupSource", "")
    evote_pickup_location = record.get("evotePickupLocation", "") or extract_pickup_location(
        evote_pickup_source,
        evote_pickup_place,
        raw_evote_pickup_rule,
        transfer_agent_name,
    )
    cached_documents = clean_pickup_documents_text(record.get("evotePickupDocuments", ""))
    parsed_documents = extract_pickup_documents(raw_evote_pickup_rule)
    evote_pickup_documents = cached_documents
    if (
        not evote_pickup_documents
        or (
            parsed_documents
            and len(parsed_documents) > len(evote_pickup_documents)
            and any(keyword in parsed_documents for keyword in ("身分證", "戶口名簿", "健保卡", "出席通知書"))
        )
    ):
        evote_pickup_documents = parsed_documents
    period_text = record.get("noticeEvotePickupPeriodText", "")
    if not period_text:
        start = record.get("evotePickupStartDate")
        end = record.get("evotePickupEndDate")
        if start and end:
            period_text = f"{start} 至 {end}"
        else:
            period_text = start or end or ""
    has_structured_snapshot_rule = bool(
        raw_evote_pickup_rule
        and period_text
        and (record.get("evotePickupStartDate") or record.get("evotePickupEndDate"))
        and re.search(r"(領取時間|領取地點|攜帶文件)[：:]", raw_evote_pickup_rule)
    )
    evote_pickup_rule = normalize_evote_rule(
        raw_evote_pickup_rule,
        period_text,
        evote_pickup_location,
        evote_pickup_documents,
    )
    if not evote_pickup_rule and has_structured_snapshot_rule:
        evote_pickup_rule = raw_evote_pickup_rule
    if not evote_pickup_rule:
        evote_pickup_location = ""
        evote_pickup_documents = ""
        period_text = ""
        record["evotePickupStartDate"] = None
        record["evotePickupEndDate"] = None

    gift_summary = trim_notice_summary(record.get("noticeGiftSummary", ""))
    notice_summary = compose_notice_summary(
        period_text,
        evote_pickup_location,
        evote_pickup_documents,
        evote_pickup_rule,
        gift_summary,
    )
    if not notice_summary and has_structured_snapshot_rule:
        notice_summary = trim_notice_summary(record.get("noticeSummary", "") or evote_pickup_rule)
    if not notice_summary and evote_pickup_rule and period_text:
        notice_summary = trim_notice_summary(evote_pickup_rule)

    record["evotePickupRule"] = evote_pickup_rule
    record["evotePickupLocation"] = evote_pickup_location
    record["evotePickupDocuments"] = evote_pickup_documents
    record["noticeEvotePickupPeriodText"] = period_text
    record["noticeGiftSummary"] = gift_summary
    record["noticeSummary"] = notice_summary
    return record


def build_notice_listing_from_meeting_date(
    code: str,
    meeting_date: str | None,
    roc_year: int | str | None = None,
) -> dict[str, str] | None:
    if not meeting_date or not re.fullmatch(r"\d{4}-\d{2}-\d{2}", meeting_date):
        return None
    normalized_year = normalize_roc_year(roc_year)
    western_year = normalized_year + 1911
    compact_meeting_date = meeting_date.replace("-", "")
    query_url = f"https://doc.twse.com.tw/server-java/t57sb01?step=1&colorchg=1&co_id={code}&year={normalized_year}&mtype=F&"
    return {
        "year": str(normalized_year),
        "dataType": "",
        "meetingType": "",
        "detail": "開會通知",
        "remark": "",
        "filename": f"{western_year}_{code}_{compact_meeting_date}F01.pdf",
        "uploadedAt": "",
        "queryUrl": query_url,
    }


def get_mops_notice_info(
    code: str,
    meeting_date: str | None = None,
    roc_year: int | str | None = None,
) -> dict[str, Any] | None:
    normalized_year = normalize_roc_year(roc_year)
    cache = load_notice_cache()
    cached_entry = notice_cache_lookup(cache, code, normalized_year)
    cached_source_type = (cached_entry or {}).get("sourceType", "")
    cached_is_current = bool(cached_entry and cached_entry.get("parserVersion") == NOTICE_CACHE_VERSION)
    cached_is_official_pdf = cached_source_type in {"official_pdf", "company_pdf", "transfer_agent_pdf"}
    if cached_entry and cached_is_current and cached_is_official_pdf:
        cached_entry["cacheStatus"] = "hit"
        return cached_entry

    listing = fetch_notice_listing(code, normalized_year)
    if not listing:
        listing = build_notice_listing_from_meeting_date(code, meeting_date, normalized_year)
    if not listing:
        return None

    cached_matches_listing = bool(cached_entry and cached_entry.get("filename") == listing["filename"])
    if (
        cached_entry
        and cached_is_current
        and cached_matches_listing
    ):
        cached_entry["cacheStatus"] = "hit"
        return cached_entry

    pdf_url = resolve_notice_pdf_url(code, listing["filename"])
    pdf_path = NOTICE_PDF_DIR / listing["filename"]
    if pdf_path.exists():
        pdf_bytes = pdf_path.read_bytes()
    else:
        pdf_bytes = fetch_bytes(pdf_url)
        pdf_path.write_bytes(pdf_bytes)

    text, text_metadata = extract_notice_text_with_metadata(pdf_bytes)
    summary = extract_notice_summary(text)
    entry = {
        "code": code,
        "rocYear": normalized_year,
        "year": normalized_year + 1911,
        "filename": listing["filename"],
        "parserVersion": NOTICE_CACHE_VERSION,
        "uploadedAt": listing["uploadedAt"],
        "queryUrl": listing["queryUrl"],
        "pdfUrl": pdf_url,
        "fetchedAt": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
        "cacheStatus": "miss",
        "textEngine": text_metadata.get("engine", ""),
        "textScore": text_metadata.get("score"),
        "ocrUsed": text_metadata.get("ocrUsed", False),
        **summary,
    }
    cache[notice_cache_storage_key(code, normalized_year)] = entry
    save_notice_cache(cache)
    return entry


def safe_get_mops_notice_info(
    code: str,
    meeting_date: str | None = None,
    roc_year: int | str | None = None,
) -> tuple[dict[str, Any] | None, str]:
    try:
        return get_mops_notice_info(code, meeting_date, roc_year), ""
    except Exception as error:
        return None, str(error)


def should_fetch_mops_notice(ideal: dict[str, Any] | None, honsec: dict[str, Any] | None) -> bool:
    has_pickup_range = bool(
        (honsec or {}).get("evote_pickup_start_date")
        or (honsec or {}).get("evote_pickup_end_date")
    )
    has_pickup_rule = bool((honsec or {}).get("evote_pickup_rule"))
    has_full_honsec = has_pickup_range and has_pickup_rule
    if has_full_honsec:
        return False

    # If ideal has no evote info at all and honsec also has nothing, MOPS is useful.
    has_any_evote = bool((ideal or {}).get("evote_start_date") or (honsec or {}).get("evote_period_text"))
    return has_any_evote or not has_full_honsec


def clean_codes(raw: str) -> list[str]:
    normalized = (
        raw.replace("（", "(")
        .replace("）", ")")
        .replace("［", "(")
        .replace("］", ")")
        .replace("【", "(")
        .replace("】", ")")
    )
    seen: set[str] = set()
    out: list[str] = []
    bracket_matches = re.findall(r"\((\d{3,6})\)", normalized)
    plain_matches = re.findall(r"\d{3,6}", normalized)
    for code in [*bracket_matches, *plain_matches]:
        if code not in seen:
            seen.add(code)
            out.append(code)
    return out


def clean_name_hint(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"^[\s\-–—:：.。]+|[\s\-–—:：.。]+$", "", text)
    text = re.sub(r"[()（）]", "", text).strip()
    if not text or text.isdigit() or len(text) > 32:
        return ""
    return text


def parse_name_hints(raw: Any) -> dict[str, str]:
    if not raw:
        return {}
    payload = raw
    if isinstance(raw, str):
        try:
            payload = json.loads(raw)
        except Exception:
            payload = raw
    hints: dict[str, str] = {}
    if isinstance(payload, dict):
        for code, name in payload.items():
            code_text = str(code)
            if re.fullmatch(r"\d{3,6}", code_text):
                cleaned = clean_name_hint(name)
                if cleaned:
                    hints[code_text] = cleaned
    elif isinstance(payload, list):
        for entry in payload:
            if isinstance(entry, dict):
                code_text = str(entry.get("code", ""))
                cleaned = clean_name_hint(entry.get("name", ""))
                if re.fullmatch(r"\d{3,6}", code_text) and cleaned:
                    hints[code_text] = cleaned
    elif isinstance(payload, str):
        normalized = (
            payload.replace("（", "(")
            .replace("）", ")")
            .replace("［", "(")
            .replace("］", ")")
            .replace("【", "(")
            .replace("】", ")")
        )
        for match in re.finditer(r"\((\d{3,6})\)", normalized):
            before = normalized[max(0, match.start() - 48) : match.start()]
            name = clean_name_hint(re.split(r"[\n\r,，、；;]+", before)[-1].split()[-1] if before.split() else "")
            if name:
                hints.setdefault(match.group(1), name)
    return hints


def apply_name_hints(results: list[dict[str, Any]], name_hints: dict[str, str]) -> list[dict[str, Any]]:
    if not name_hints:
        return results
    for item in results:
        code = str(item.get("code", ""))
        hint = name_hints.get(code, "")
        if hint and not item.get("companyName"):
            item["companyName"] = hint
            item["inputCompanyName"] = hint
    return results


def apply_name_hints_to_compare_rows(rows: list[dict[str, Any]], name_hints: dict[str, str]) -> list[dict[str, Any]]:
    if not name_hints:
        return rows
    for row in rows:
        code = str(row.get("code", ""))
        hint = name_hints.get(code, "")
        if hint and not row.get("companyName"):
            row["companyName"] = hint
            row["inputCompanyName"] = hint
        for item in row.get("years", []):
            if isinstance(item, dict) and hint and not item.get("companyName"):
                item["companyName"] = hint
                item["inputCompanyName"] = hint
    return rows


def source_link(label: str, url: str) -> dict[str, str]:
    return {"label": label, "url": url}


def item_has_pickup_info(item: dict[str, Any]) -> bool:
    return bool(
        item.get("evotePickupStartDate")
        or item.get("evotePickupEndDate")
        or item.get("noticeEvotePickupPeriodText")
        or item.get("evotePickupRule")
    )


def item_has_notice_info(item: dict[str, Any]) -> bool:
    return bool(item.get("noticeFilename") or item.get("noticeSourceLabel") or item.get("noticeSummary"))


def missing_data_reason(item: dict[str, Any]) -> str:
    status = item.get("status", "")
    has_notice = item_has_notice_info(item)
    has_pickup = item_has_pickup_info(item)
    note = item.get("notes", "")
    mops_error = item.get("mopsError", "")
    cache_status = item.get("noticeCacheStatus", "")

    if status == "published":
        if has_pickup:
            return ""
        if has_notice:
            return "已公告紀念品，且已有通知書快取，但尚未解析出電子投票股東領取日期/地點。"
        return "已公告紀念品，但尚未取得通知書或電子投票股東領取資訊。"

    if status == "partial":
        if has_pickup:
            return "已有部分會議/通知書資料，但紀念品公告來源仍未完整。"
        if has_notice:
            return "已有部分會議/通知書資料，但尚未解析出電子投票股東領取資訊。"
        return note or "只有部分資料，尚未補齊紀念品與通知書。"

    if has_notice:
        summary = item.get("noticeSummary", "") or item.get("noticeGiftSummary", "")
        if summary:
            return "紀念品來源未收錄；已有通知書摘要，請人工確認是否未發紀念品或解析仍不足。"
        return "紀念品來源未收錄；已有通知書快取，但未解析出紀念品或電子投票領取資訊。"
    if cache_status == "miss":
        return "紀念品來源未收錄；MOPS/官方通知書查詢未命中。"
    if mops_error:
        return f"紀念品來源未收錄；通知書查詢錯誤：{mops_error}"
    return note or "紀念品來源未收錄或今年尚未公告。"


def build_export_rows(results: list[dict[str, Any]]) -> list[list[str]]:
    rows = [[
        "股票代號",
        "公司名稱",
        "狀態",
        "查無/缺資料原因",
        "紀念品",
        "最後買進日",
        "股東會日期",
        "股東會地點",
        "電子投票開始",
        "電子投票結束",
        "電投領取開始",
        "電投領取結束",
        "電投領取期間原文",
        "電投領取來源",
        "電投領取地點",
        "電投攜帶資料",
        "電投領取資訊",
        "通知書摘要(電投重點)",
        "通知書快取",
        "股代名稱",
        "股代電話",
        "零股寄單",
        "資料來源",
    ]]
    for item in results:
        rows.append([
            item.get("code", ""),
            item.get("companyName", ""),
            item.get("status", ""),
            missing_data_reason(item),
            item.get("souvenirName", ""),
            item.get("lastBuyDate", "") or "",
            item.get("meetingDate", "") or "",
            item.get("meetingCity", ""),
            item.get("evoteStartDate", "") or "",
            item.get("evoteEndDate", "") or "",
            item.get("evotePickupStartDate", "") or "",
            item.get("evotePickupEndDate", "") or "",
            item.get("noticeEvotePickupPeriodText", ""),
            item.get("evotePickupSource", ""),
            item.get("evotePickupLocation", ""),
            item.get("evotePickupDocuments", ""),
            item.get("evotePickupRule", ""),
            item.get("noticeSummary", "") or item.get("noticeGiftSummary", ""),
            item.get("noticeCacheStatus", ""),
            item.get("transferAgentName", "") or item.get("transferAgentShort", ""),
            item.get("transferAgentPhone", ""),
            item.get("oddLotMail", ""),
            " | ".join(source.get("label", "") for source in item.get("sources", [])),
        ])
    return rows


def export_row_values(item: dict[str, Any]) -> list[str]:
    return [
        item.get("code", ""),
        item.get("companyName", ""),
        item.get("status", ""),
        missing_data_reason(item),
        item.get("souvenirName", ""),
        item.get("lastBuyDate", "") or "",
        item.get("meetingDate", "") or "",
        item.get("meetingCity", ""),
        item.get("evoteStartDate", "") or "",
        item.get("evoteEndDate", "") or "",
        item.get("evotePickupStartDate", "") or "",
        item.get("evotePickupEndDate", "") or "",
        item.get("noticeEvotePickupPeriodText", ""),
        item.get("evotePickupSource", ""),
        item.get("evotePickupLocation", ""),
        item.get("evotePickupDocuments", ""),
        item.get("evotePickupRule", ""),
        item.get("noticeSummary", "") or item.get("noticeGiftSummary", ""),
        item.get("noticeCacheStatus", ""),
        item.get("transferAgentName", "") or item.get("transferAgentShort", ""),
        item.get("transferAgentPhone", ""),
        item.get("oddLotMail", ""),
        " | ".join(source.get("label", "") for source in item.get("sources", [])),
    ]


def build_export_xlsx(results: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "股東會紀念品"

    for row in build_export_rows(results):
        worksheet.append([excel_safe_value(value) for value in row])

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    wrap_columns = {"D", "E", "M", "O", "P", "Q", "R", "W"}
    widths = {
        "A": 12, "B": 18, "C": 10, "D": 38, "E": 28, "F": 14, "G": 14,
        "H": 12, "I": 14, "J": 14, "K": 14, "L": 14, "M": 24, "N": 12,
        "O": 24, "P": 36, "Q": 48, "R": 56, "S": 12, "T": 22, "U": 18,
        "V": 10, "W": 24,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column_letter in wrap_columns)
            if isinstance(cell.value, date):
                cell.number_format = "yyyy-mm-dd"

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_compare_export_xlsx(compare_rows: list[dict[str, Any]], years: list[int]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "三年度比較"

    year_fields = [
        ("狀態", "status"),
        ("查無/缺資料原因", "_missingReason"),
        ("紀念品", "souvenirName"),
        ("最後買進日", "lastBuyDate"),
        ("股東會日期", "meetingDate"),
        ("股東會地點", "meetingCity"),
        ("電子投票開始", "evoteStartDate"),
        ("電子投票結束", "evoteEndDate"),
        ("電投領取開始", "evotePickupStartDate"),
        ("電投領取結束", "evotePickupEndDate"),
        ("電投領取期間原文", "noticeEvotePickupPeriodText"),
        ("電投領取來源", "evotePickupSource"),
        ("電投領取地點", "evotePickupLocation"),
        ("電投攜帶資料", "evotePickupDocuments"),
        ("電投領取資訊", "evotePickupRule"),
        ("通知書摘要(電投重點)", "noticeSummary"),
        ("通知書來源", "noticeSourceLabel"),
        ("股代名稱", "transferAgentName"),
    ]
    header = ["股票代號", "公司名稱"]
    for year in years:
        header.extend(f"{year}年 {label}" for label, _ in year_fields)
    worksheet.append(header)

    for row in compare_rows:
        history_by_year = {
            int(item.get("rocYear") or 0): item
            for item in row.get("years", [])
            if isinstance(item, dict)
        }
        values = [row.get("code", ""), row.get("companyName", "")]
        for year in years:
            item = history_by_year.get(year, {})
            for _, field in year_fields:
                if field == "_missingReason":
                    values.append(missing_data_reason(item))
                elif field == "noticeSummary":
                    values.append(item.get("noticeSummary", "") or item.get("noticeGiftSummary", ""))
                elif field == "transferAgentName":
                    values.append(item.get("transferAgentName", "") or item.get("transferAgentShort", ""))
                else:
                    values.append(item.get(field, "") or "")
        worksheet.append([excel_safe_value(value) for value in values])

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    worksheet.freeze_panes = "C2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.column_dimensions["A"].width = 12
    worksheet.column_dimensions["B"].width = 18
    for column_cells in worksheet.iter_cols(min_col=3, max_col=worksheet.max_column):
        column_letter = column_cells[0].column_letter
        worksheet.column_dimensions[column_letter].width = 18

    wide_wrap_labels = {"原因", "領取地點", "攜帶資料", "領取資訊", "通知書摘要", "期間原文"}
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            header_value = str(worksheet.cell(row=1, column=cell.column).value or "")
            should_wrap = any(label in header_value for label in wide_wrap_labels)
            cell.alignment = Alignment(vertical="top", wrap_text=should_wrap)
            if isinstance(cell.value, date):
                cell.number_format = "yyyy-mm-dd"

    detail_sheet = workbook.create_sheet("年度明細")
    detail_header = ["年度", *build_export_rows([])[0]]
    detail_sheet.append(detail_header)
    for row in compare_rows:
        for item in row.get("years", []):
            detail_sheet.append([excel_safe_value(item.get("rocYear", "")), *[excel_safe_value(value) for value in export_row_values(item)]])

    for cell in detail_sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    detail_sheet.freeze_panes = "A2"
    detail_sheet.auto_filter.ref = detail_sheet.dimensions
    detail_widths = {
        "A": 10, "B": 12, "C": 18, "D": 10, "E": 38, "F": 28, "G": 14,
        "H": 14, "I": 12, "J": 14, "K": 14, "L": 14, "M": 14, "N": 24,
        "O": 12, "P": 24, "Q": 36, "R": 48, "S": 56, "T": 12, "U": 22,
        "V": 18, "W": 10, "X": 24,
    }
    for column, width in detail_widths.items():
        detail_sheet.column_dimensions[column].width = width
    for row in detail_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column_letter in {"E", "F", "N", "P", "Q", "R", "S", "X"})
            if isinstance(cell.value, date):
                cell.number_format = "yyyy-mm-dd"

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def excel_safe_value(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    cleaned = EXCEL_ILLEGAL_CHAR_RE.sub("", value)[:32767]
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", cleaned):
        try:
            return date.fromisoformat(cleaned)
        except ValueError:
            return cleaned
    return cleaned


def empty_record(code: str, note: str = "") -> dict[str, Any]:
    return {
        "code": code,
        "rocYear": CURRENT_ROC_YEAR,
        "year": CURRENT_YEAR,
        "companyName": "",
        "status": "unpublished",
        "isPublished": False,
        "souvenirName": "",
        "meetingDate": "",
        "lastBuyDate": "",
        "meetingCity": "",
        "priceText": "",
        "transferAgentName": "",
        "transferAgentPhone": "",
        "transferAgentShort": "",
        "oddLotMail": "",
        "reelection": "",
        "needVote": None,
        "fractionalOk": None,
        "evoteStartDate": None,
        "evoteEndDate": None,
        "evotePickupStartDate": None,
        "evotePickupEndDate": None,
        "evotePickupPlace": "",
        "evotePickupLocation": "",
        "evotePickupDocuments": "",
        "evotePickupRule": "",
        "meetingDistributionRule": "",
        "proxyPeriodText": "",
        "agentDistributionPeriodText": "",
        "noticeSummary": "",
        "noticeGiftSummary": "",
        "noticeFilename": "",
        "noticeUploadedAt": "",
        "noticeEvotePickupPeriodText": "",
        "noticeCacheStatus": "",
        "noticeSourceLabel": "",
        "noticeSourceType": "",
        "mopsAttempted": False,
        "mopsError": "",
        "evotePickupSource": "",
        "notes": note,
        "sources": [],
    }


def build_lookup_results(
    codes: list[str],
    *,
    allow_live_lookup: bool | None = None,
    allow_live_notice_fetch: bool | None = None,
    roc_year: int | str | None = None,
    snapshot_only: bool = False,
) -> tuple[list[dict[str, Any]], dict[str, int], dict[str, Any]]:
    normalized_year = normalize_roc_year(roc_year)
    snapshot = load_lookup_snapshot(normalized_year)
    snapshot_records = snapshot.get("records") if isinstance(snapshot.get("records"), dict) else {}
    use_snapshot = bool(snapshot_records)
    effective_live_lookup = ALLOW_LIVE_LOOKUP if allow_live_lookup is None else allow_live_lookup
    if allow_live_notice_fetch is None:
        effective_notice_fetch = effective_live_lookup or not use_snapshot
    else:
        effective_notice_fetch = allow_live_notice_fetch

    results_map: dict[str, dict[str, Any]] = {}
    missing_codes: list[str] = []
    notice_cache = load_notice_cache() if use_snapshot else {}
    for code in codes:
        record = snapshot_records.get(code) if use_snapshot else None
        if isinstance(record, dict):
            cached_notice = notice_cache_lookup(notice_cache, code, normalized_year)
            merged_record = merge_notice_cache_into_record(dict(record), cached_notice)
            results_map[code] = enrich_record_notice_fields(merged_record)
        else:
            missing_codes.append(code)

    source_stats = snapshot.get("sourceStats", {}) if use_snapshot else {}
    if missing_codes and not snapshot_only and (effective_live_lookup or not use_snapshot):
        sources = source_bundle(normalized_year)
        for code in missing_codes:
            results_map[code] = build_record(
                code,
                sources,
                allow_live_notice_fetch=effective_notice_fetch,
                roc_year=normalized_year,
            )
        if not source_stats:
            source_stats = build_source_stats(sources)
    else:
        for code in missing_codes:
            record = empty_record(code, "這檔資料尚未由本機同步流程補齊，請先更新本機資料後再部署。")
            record["rocYear"] = normalized_year
            record["year"] = normalized_year + 1911
            results_map[code] = record

    if not isinstance(source_stats, dict):
        source_stats = {}
    normalized_source_stats = {
        "wespai": int(source_stats.get("wespai", 0)),
        "idealLabs": int(source_stats.get("idealLabs", 0)),
        "honsec": int(source_stats.get("honsec", 0)),
    }
    metadata = snapshot_metadata(normalized_year)
    results = [results_map.get(code, empty_record(code)) for code in codes]
    return results, normalized_source_stats, metadata


def build_record(
    code: str,
    sources: dict[str, Any],
    allow_live_notice_fetch: bool = True,
    roc_year: int | str | None = None,
) -> dict[str, Any]:
    normalized_year = normalize_roc_year(roc_year)
    wespai = sources["wespai"].get(code)
    ideal = sources["ideal"].get(code)
    honsec = sources["honsec"].get(code)
    company_name = (
        (wespai or {}).get("company_name")
        or (ideal or {}).get("stock", {}).get("stock_name")
        or (honsec or {}).get("company_name")
        or ""
    )
    meeting_date = (ideal or {}).get("meeting_date") or (wespai or {}).get("meeting_date")
    last_buy_date = (ideal or {}).get("last_buy_date") or (wespai or {}).get("last_buy_date") or (honsec or {}).get(
        "last_buy_date"
    )
    should_try_mops = bool((wespai or ideal or honsec) and should_fetch_mops_notice(ideal, honsec))
    mops_notice = None
    mops_error = ""
    cached_notice = notice_cache_lookup(load_notice_cache(), code, normalized_year)
    if should_try_mops:
        if allow_live_notice_fetch:
            mops_notice, mops_error = safe_get_mops_notice_info(code, meeting_date, normalized_year)
        else:
            if isinstance(cached_notice, dict):
                mops_notice = cached_notice
    elif isinstance(cached_notice, dict):
        mops_notice = cached_notice

    evote_start = (ideal or {}).get("evote_start_date") or (honsec or {}).get("evote_start_date")
    evote_end = (ideal or {}).get("evote_end_date") or (honsec or {}).get("evote_end_date")
    pickup_start = (honsec or {}).get("evote_pickup_start_date") or (mops_notice or {}).get("evotePickupStartDate")
    pickup_end = (honsec or {}).get("evote_pickup_end_date") or (mops_notice or {}).get("evotePickupEndDate")
    souvenir_name = (
        (wespai or {}).get("souvenir_name")
        or (ideal or {}).get("souvenir_name")
        or (honsec or {}).get("souvenir_name")
        or ""
    )
    transfer_agent_name = (
        (ideal or {}).get("transfer_agent_name")
        or (wespai or {}).get("transfer_agent_short")
        or ""
    )
    transfer_agent_phone = (
        (ideal or {}).get("transfer_agent_phone")
        or (wespai or {}).get("transfer_agent_phone")
        or ""
    )

    source_links: list[dict[str, str]] = []
    if wespai:
        source_links.append(source_link("撿股讚", wespai["source_url"]))
        source_links.append(source_link("官方開會資料", wespai["official_doc_url"]))
    if ideal:
        source_links.append(source_link("股東禮簿", IDEAL_URL))
    if honsec:
        source_links.append(source_link("宏遠股代", HONSEC_URL))
    if mops_notice:
        notice_label = (mops_notice.get("sourceLabel") or "開會通知書")
        source_links.append(source_link(notice_label, mops_notice["pdfUrl"]))

    is_published = bool(wespai or ideal or honsec and honsec.get("souvenir_name"))
    if not is_published and company_name:
        status = "partial"
    elif is_published:
        status = "published"
    else:
        status = "unpublished"

    mops_has_pickup_info = bool(
        (mops_notice or {}).get("evotePickupRule")
        or (mops_notice or {}).get("evotePickupStartDate")
        or (mops_notice or {}).get("evotePickupEndDate")
        or (mops_notice or {}).get("evotePickupPeriodText")
    )
    evote_pickup_source = (
        "宏遠股代"
        if (honsec or {}).get("evote_pickup_rule")
        else ("開會通知書" if mops_has_pickup_info else "")
    )
    raw_evote_pickup_rule = (honsec or {}).get("evote_pickup_rule") or (mops_notice or {}).get("evotePickupRule", "")
    evote_pickup_location = extract_pickup_location(
        evote_pickup_source,
        (honsec or {}).get("evote_pickup_place", ""),
        raw_evote_pickup_rule,
        transfer_agent_name,
    )
    evote_pickup_location = (mops_notice or {}).get("evotePickupLocation") or evote_pickup_location
    evote_pickup_documents = (mops_notice or {}).get("evotePickupDocuments") or extract_pickup_documents(
        raw_evote_pickup_rule
    )
    evote_pickup_period_text = (mops_notice or {}).get("evotePickupPeriodText", "")
    if not evote_pickup_period_text:
        if pickup_start and pickup_end:
            evote_pickup_period_text = f"{pickup_start} 至 {pickup_end}"
        else:
            evote_pickup_period_text = pickup_start or pickup_end or ""
    evote_pickup_rule = normalize_evote_rule(
        raw_evote_pickup_rule,
        evote_pickup_period_text,
        evote_pickup_location,
        evote_pickup_documents,
    )
    if not evote_pickup_rule:
        pickup_start = None
        pickup_end = None
        evote_pickup_location = ""
        evote_pickup_documents = ""
        evote_pickup_period_text = ""
        evote_pickup_source = ""
    notice_summary = compose_notice_summary(
        evote_pickup_period_text,
        evote_pickup_location,
        evote_pickup_documents,
        raw_evote_pickup_rule,
        (mops_notice or {}).get("giftSummary", ""),
    )

    return {
        "code": code,
        "rocYear": normalized_year,
        "year": normalized_year + 1911,
        "companyName": company_name,
        "status": status,
        "isPublished": is_published,
        "souvenirName": souvenir_name,
        "meetingDate": meeting_date,
        "lastBuyDate": last_buy_date,
        "meetingCity": (wespai or {}).get("meeting_city", ""),
        "priceText": (wespai or {}).get("price_text", ""),
        "transferAgentName": transfer_agent_name,
        "transferAgentPhone": transfer_agent_phone,
        "transferAgentShort": (wespai or {}).get("transfer_agent_short", ""),
        "oddLotMail": (wespai or {}).get("odd_lot_mail", ""),
        "reelection": (wespai or {}).get("reelection", ""),
        "needVote": (ideal or {}).get("need_vote"),
        "fractionalOk": (ideal or {}).get("fractional_ok"),
        "evoteStartDate": evote_start,
        "evoteEndDate": evote_end,
        "evotePickupStartDate": pickup_start,
        "evotePickupEndDate": pickup_end,
        "evotePickupPlace": (honsec or {}).get("evote_pickup_place", ""),
        "evotePickupLocation": evote_pickup_location,
        "evotePickupDocuments": evote_pickup_documents,
        "evotePickupRule": evote_pickup_rule,
        "meetingDistributionRule": (honsec or {}).get("meeting_distribution_rule", ""),
        "proxyPeriodText": (honsec or {}).get("proxy_period_text", ""),
        "agentDistributionPeriodText": (honsec or {}).get("agent_distribution_period_text", ""),
        "noticeSummary": notice_summary,
        "noticeGiftSummary": (mops_notice or {}).get("giftSummary", ""),
        "noticeFilename": (mops_notice or {}).get("filename", ""),
        "noticeUploadedAt": (mops_notice or {}).get("uploadedAt", ""),
        "noticeEvotePickupPeriodText": evote_pickup_period_text,
        "noticeCacheStatus": (mops_notice or {}).get("cacheStatus", ""),
        "noticeSourceLabel": (mops_notice or {}).get("sourceLabel", "公開資訊觀測站通知書" if mops_notice else ""),
        "noticeSourceType": (mops_notice or {}).get("sourceType", "mops" if mops_notice else ""),
        "mopsAttempted": bool(mops_notice or (should_try_mops and allow_live_notice_fetch)),
        "mopsError": mops_error,
        "evotePickupSource": evote_pickup_source,
        "notes": (
            "目前在整合來源中尚未看到該年度紀念品公告，建議保留 watchlist 持續追蹤。"
            if status == "unpublished"
            else ""
        ),
        "sources": source_links,
    }


def build_compare_results(
    codes: list[str],
    roc_years: list[int] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    years = roc_years or DEFAULT_COMPARE_ROC_YEARS
    per_year_records: dict[int, dict[str, dict[str, Any]]] = {}
    per_year_metadata: dict[str, Any] = {}
    for roc_year in years:
        results, source_stats, metadata = build_lookup_results(
            codes,
            roc_year=roc_year,
            allow_live_lookup=False,
            allow_live_notice_fetch=False,
            snapshot_only=True,
        )
        per_year_records[roc_year] = {item["code"]: item for item in results}
        per_year_metadata[str(roc_year)] = {
            **metadata,
            "sourceStats": source_stats,
        }

    rows: list[dict[str, Any]] = []
    for code in codes:
        history = [per_year_records.get(roc_year, {}).get(code, empty_record(code)) for roc_year in years]
        company_name = next((item.get("companyName") for item in history if item.get("companyName")), "")
        rows.append(
            {
                "code": code,
                "companyName": company_name,
                "years": history,
            }
        )

    return rows, {
        "years": years,
        "perYear": per_year_metadata,
    }


class AppHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs) -> None:
        super().__init__(*args, directory=str(ROOT), **kwargs)

    def do_POST(self) -> None:
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path == "/api/export.xlsx":
            self.handle_export_post()
            return
        json_response(self, {"ok": False, "error": "Not found"}, status=404)

    def do_GET(self) -> None:
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path == "/api/lookup":
            query = urllib.parse.parse_qs(parsed.query)
            raw_codes = query.get("codes", [""])[0]
            raw_years = query.get("years", [""])[0]
            raw_names = query.get("names", [""])[0]
            self.handle_lookup(raw_codes, raw_years, raw_names=raw_names)
            return
        if parsed.path == "/api/compare":
            query = urllib.parse.parse_qs(parsed.query)
            raw_codes = query.get("codes", [""])[0]
            raw_years = query.get("years", [""])[0]
            raw_names = query.get("names", [""])[0]
            self.handle_compare(raw_codes, raw_years, raw_names=raw_names)
            return
        if parsed.path == "/api/export.xlsx":
            query = urllib.parse.parse_qs(parsed.query)
            raw_codes = query.get("codes", [""])[0]
            raw_years = query.get("years", [""])[0]
            mode = query.get("mode", [""])[0]
            raw_names = query.get("names", [""])[0]
            self.handle_export(raw_codes, raw_years=raw_years, mode=mode, raw_names=raw_names)
            return
        if parsed.path == "/api/health":
            json_response(
                self,
                {
                    "ok": True,
                    "generatedAt": date.today().isoformat(),
                    "version": APP_VERSION,
                    **snapshot_metadata(),
                },
            )
            return
        if parsed.path == "/api/notice-progress":
            json_response(
                self,
                {
                    "ok": True,
                    "generatedAt": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
                    **build_notice_progress(),
                },
            )
            return
        if parsed.path == "/api/requested-codes":
            requested = load_requested_codes()
            json_response(
                self,
                {
                    "ok": True,
                    "generatedAt": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
                    "codes": sorted(requested),
                    "requested": requested,
                },
            )
            return
        if parsed.path == "/":
            self.path = "/index.html"
        super().do_GET()

    def handle_lookup(self, raw_codes: str, raw_years: str = "", raw_names: Any = "") -> None:
        codes = clean_codes(raw_codes)
        if not codes:
            json_response(
                self,
                {
                    "ok": False,
                    "error": "請輸入至少一筆股票代號。",
                    "results": [],
                },
                status=400,
            )
            return

        try:
            record_requested_codes(codes)
            roc_years = parse_roc_years(raw_years)
            results, source_stats, metadata = build_lookup_results(codes, roc_year=roc_years[0])
            apply_name_hints(results, parse_name_hints(raw_names))
            json_response(
                self,
                {
                    "ok": True,
                    "generatedAt": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
                    "requestedCodes": codes,
                    "requestedYears": roc_years,
                    "sourceStats": source_stats,
                    **metadata,
                    "results": results,
                },
            )
        except Exception as error:  # pragma: no cover - we want a readable API error
            json_response(
                self,
                {
                    "ok": False,
                    "error": f"資料抓取失敗：{error}",
                    "results": [],
                },
                status=502,
            )

    def handle_compare(self, raw_codes: str, raw_years: str = "", raw_names: Any = "") -> None:
        codes = clean_codes(raw_codes)
        if not codes:
            json_response(
                self,
                {
                    "ok": False,
                    "error": "請輸入至少一筆股票代號。",
                    "results": [],
                },
                status=400,
            )
            return

        try:
            record_requested_codes(codes)
            years = parse_roc_years(raw_years) if raw_years else DEFAULT_COMPARE_ROC_YEARS
            results, metadata = build_compare_results(codes, years)
            apply_name_hints_to_compare_rows(results, parse_name_hints(raw_names))
            json_response(
                self,
                {
                    "ok": True,
                    "generatedAt": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
                    "requestedCodes": codes,
                    **metadata,
                    "results": results,
                },
            )
        except Exception as error:
            json_response(
                self,
                {
                    "ok": False,
                    "error": f"三年度比較資料讀取失敗：{error}",
                    "results": [],
                },
                status=502,
            )

    def handle_export(self, raw_codes: str, raw_years: str = "", mode: str = "", raw_names: Any = "") -> None:
        codes = clean_codes(raw_codes)
        if not codes:
            json_response(
                self,
                {"ok": False, "error": "請輸入至少一筆股票代號。"},
                status=400,
            )
            return
        try:
            export_mode = (mode or "").lower()
            name_hints = parse_name_hints(raw_names)
            if export_mode == "compare":
                years = parse_roc_years(raw_years) if raw_years else DEFAULT_COMPARE_ROC_YEARS
                compare_rows, _ = build_compare_results(codes, years)
                apply_name_hints_to_compare_rows(compare_rows, name_hints)
                body = build_compare_export_xlsx(compare_rows, years)
                year_part = "-".join(str(year) for year in years)
                filename = f"shareholder-gifts-compare-{year_part}-{time.strftime('%Y%m%d-%H%M%S')}.xlsx"
            else:
                results, _, _ = build_lookup_results(codes)
                apply_name_hints(results, name_hints)
                body = build_export_xlsx(results)
                filename = f"shareholder-gifts-{time.strftime('%Y%m%d-%H%M%S')}.xlsx"
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
            self.send_header("Content-Length", str(len(body)))
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            self.wfile.write(body)
        except Exception as error:  # pragma: no cover
            json_response(
                self,
                {"ok": False, "error": f"Excel 匯出失敗：{error}"},
                status=502,
            )

    def handle_export_post(self) -> None:
        try:
            length = int(self.headers.get("Content-Length", "0"))
        except ValueError:
            length = 0
        raw_body = self.rfile.read(max(0, length))
        content_type = (self.headers.get("Content-Type") or "").lower()
        codes: list[str] = []
        mode = ""
        raw_years = ""
        raw_names: Any = ""
        if "application/json" in content_type:
            try:
                payload = json.loads(raw_body.decode("utf-8")) if raw_body else {}
            except Exception:
                json_response(self, {"ok": False, "error": "匯出參數格式錯誤。"}, status=400)
                return
            codes = clean_codes(" ".join(str(code) for code in payload.get("codes", [])))
            mode = str(payload.get("mode", ""))
            raw_years = str(payload.get("years", ""))
            raw_names = payload.get("names", "")
        else:
            try:
                parsed = urllib.parse.parse_qs(raw_body.decode("utf-8"))
            except Exception:
                json_response(self, {"ok": False, "error": "匯出參數格式錯誤。"}, status=400)
                return
            raw_codes = parsed.get("codes", [""])[0]
            codes = clean_codes(raw_codes)
            mode = parsed.get("mode", [""])[0]
            raw_years = parsed.get("years", [""])[0]
            raw_names = parsed.get("names", [""])[0]
        self.handle_export(",".join(codes), raw_years=raw_years, mode=mode, raw_names=raw_names)


def main() -> None:
    server = ThreadingHTTPServer((HOST, PORT), AppHandler)
    print(f"Shareholder gift tracker running at http://127.0.0.1:{PORT}")
    server.serve_forever()


if __name__ == "__main__":
    main()
